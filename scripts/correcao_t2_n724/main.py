from __future__ import annotations

import argparse
import csv
import html
import json
import math
import re
import unicodedata
from io import BytesIO
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
import pdfplumber


URL_RE = re.compile(r"https?://[^\s\"'<>]+")
SPREADSHEET_ID_RE = re.compile(r"/spreadsheets/d/([A-Za-z0-9_-]+)")


@dataclass
class PageExtraction:
    page: int
    chars: int
    words: int
    tables: int
    likely_image: bool


@dataclass
class PdfExtractionReport:
    file: str
    pages: int
    chars: int
    words: int
    tables: int
    likely_scanned: bool
    output_md: str
    output_txt: str
    page_reports: list[PageExtraction]


@dataclass
class LinkBundle:
    forms: str = ""
    csv: str = ""
    sheet: str = ""
    all_urls: list[str] | None = None


@dataclass
class RubricItem:
    item: str
    max_score: float
    score: float
    status: str
    evidence: str
    flags: list[str]


@dataclass
class GroupAudit:
    group: str
    ava_file: str
    feedback_file: str
    dreamshaper_file: str
    dreamshaper_md: str
    links: LinkBundle
    rubric: list[RubricItem]
    flags: list[str]
    notes: list[str]
    local_score: float
    local_score_rounded: float
    final_status: str


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="python -m scripts.correcao_t2_n724")
    sub = parser.add_subparsers(dest="cmd", required=True)

    extract = sub.add_parser("extract-dreamshaper-pdf", help="Extrai PDFs do DreamShaper para MD/TXT auditaveis")
    extract.add_argument(
        "--input",
        type=Path,
        default=Path("trabalhos/N724/und2/correcao_T2/entradas/dreamshaper"),
        help="Arquivo PDF ou diretorio com PDFs",
    )
    extract.add_argument(
        "--out",
        type=Path,
        default=Path("trabalhos/N724/und2/correcao_T2/_saida/dreamshaper_extraido"),
        help="Diretorio de saida",
    )
    extract.add_argument("--min-chars-page", type=int, default=80)

    run_p = sub.add_parser("run", help="Executa a auditoria local do T2 N724")
    run_p.add_argument(
        "--base",
        type=Path,
        default=Path("trabalhos/N724/und2/correcao_T2"),
        help="Diretorio da correcao_T2",
    )
    run_p.add_argument(
        "--out",
        type=Path,
        default=Path("trabalhos/N724/und2/correcao_T2/_saida"),
        help="Diretorio de saida",
    )
    run_p.add_argument("--check-remote", action="store_true", help="Tenta baixar e auditar links externos")

    args = parser.parse_args(argv)
    if args.cmd == "extract-dreamshaper-pdf":
        return extract_dreamshaper_pdf(args.input, args.out, args.min_chars_page)
    if args.cmd == "run":
        return run_pipeline(args.base, args.out, args.check_remote)
    return 2


def run_pipeline(base: Path, out: Path, check_remote: bool) -> int:
    required = {
        "ava": base / "entradas" / "ava_moodle",
        "dreamshaper": base / "entradas" / "dreamshaper",
        "feedbacks": base / "entradas" / "feedbacks_T1",
        "dreamshaper_extraido": out / "dreamshaper_extraido",
        "rubrica": base / "referencias" / "rubrica_operacional.md",
    }
    missing = [str(path) for path in required.values() if not path.exists()]
    if missing:
        print("Entradas obrigatorias ausentes:")
        for path in missing:
            print(f"- {path}")
        return 2

    ensure_output_dirs(out)
    audits = build_group_audits(base, out)
    if check_remote:
        apply_remote_audit(audits, out)
    write_pipeline_outputs(audits, out)
    return 0


def ensure_output_dirs(out: Path) -> None:
    for path in [out, out / "rubricas", out / "downloads", out / "auditoria_estatistica", out / "auditoria_interpretativa"]:
        path.mkdir(parents=True, exist_ok=True)


def build_group_audits(base: Path, out: Path) -> list[GroupAudit]:
    ava_files = sorted((base / "entradas" / "ava_moodle").glob("Grupo_*/textoonline.html"))
    feedback_files = {
        "Grupo A": base / "entradas" / "feedbacks_T1" / "grupoA_feedbacks.md",
        "Grupo B": base / "entradas" / "feedbacks_T1" / "grupoB_feedbacks.md",
    }
    dream_docs = load_dreamshaper_docs(out / "dreamshaper_extraido")
    audits: list[GroupAudit] = []

    for ava_file in ava_files:
        group = normalize_group_from_path(ava_file)
        feedback_file = feedback_files.get(group, Path())
        feedback_text = read_text(feedback_file)
        dream_md, dream_text = match_dreamshaper_doc(group, feedback_text, dream_docs)
        links = extract_links_from_ava(ava_file)
        rubric, flags, notes = evaluate_group(group, links, feedback_text, dream_text)
        rubric = round_rubric_items(rubric)
        local_score = sum(item.score for item in rubric)
        audits.append(
            GroupAudit(
                group=group,
                ava_file=str(ava_file),
                feedback_file=str(feedback_file),
                dreamshaper_file=find_original_pdf(base, dream_md),
                dreamshaper_md=str(dream_md),
                links=links,
                rubric=rubric,
                flags=sorted(set(flags)),
                notes=notes,
                local_score=local_score,
                local_score_rounded=round_up_to_step(local_score, 0.05),
                final_status="preliminar_revisao_manual",
            )
        )
    return audits


def load_dreamshaper_docs(root: Path) -> list[tuple[Path, str]]:
    docs = []
    for path in sorted(root.glob("*.md")):
        if path.name.startswith("dreamshaper_pdf_diagnostico"):
            continue
        docs.append((path, read_text(path)))
    return docs


def normalize_group_from_path(path: Path) -> str:
    name = path.parent.name.replace("_", " ")
    match = re.search(r"Grupo\s+([A-Z])", name, re.IGNORECASE)
    return f"Grupo {match.group(1).upper()}" if match else name


def read_text(path: Path) -> str:
    if not path or not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def match_dreamshaper_doc(group: str, feedback_text: str, docs: list[tuple[Path, str]]) -> tuple[Path, str]:
    feedback_tokens = content_tokens(feedback_text)
    best: tuple[Path, str] | None = None
    best_score = -1
    for path, text in docs:
        score = len(feedback_tokens & content_tokens(text))
        score += project_match_bonus(group, feedback_text, text, path)
        if score > best_score:
            best = (path, text)
            best_score = score
    return best if best is not None else (Path(), "")


def project_match_bonus(group: str, feedback_text: str, dream_text: str, path: Path) -> int:
    combined = f"{feedback_text}\n{dream_text}\n{path.name}".lower()
    dream_lower = dream_text.lower()
    bonus = 0
    if group == "Grupo A" and "direito" in dream_lower and ("computação" in dream_lower or "computacao" in dream_lower):
        bonus += 100
    if group == "Grupo B" and "classe social" in dream_lower:
        bonus += 100
    if "classe social" in feedback_text.lower() and "classe social" in dream_lower:
        bonus += 50
    if "direito" in feedback_text.lower() and "direito" in dream_lower and ("computação" in combined or "computacao" in combined):
        bonus += 50
    return bonus


def content_tokens(text: str) -> set[str]:
    stop = {
        "para",
        "com",
        "que",
        "uma",
        "dos",
        "das",
        "por",
        "mais",
        "como",
        "esta",
        "este",
        "isso",
        "deve",
        "ainda",
        "grupo",
        "projeto",
        "pesquisa",
    }
    return {t for t in re.findall(r"[A-Za-zÀ-ÿ0-9]{4,}", text.lower()) if t not in stop}


def extract_links_from_ava(path: Path) -> LinkBundle:
    raw = read_text(path)
    soup = BeautifulSoup(raw, "html.parser")
    lines = [line.strip() for line in soup.get_text("\n").splitlines() if line.strip()]
    urls = [clean_url(url) for url in URL_RE.findall(html.unescape(raw))]
    bundle = LinkBundle(all_urls=unique(urls))

    for idx, line in enumerate(lines):
        lower = line.lower()
        next_url = first_url(line) or first_url(lines[idx + 1] if idx + 1 < len(lines) else "")
        if not next_url:
            continue
        if "forms" in lower or "formul" in lower:
            bundle.forms = clean_url(next_url)
        elif "csv" in lower:
            bundle.csv = clean_url(next_url)
        elif "planilha" in lower:
            bundle.sheet = clean_url(next_url)

    if not bundle.forms:
        bundle.forms = next((u for u in urls if "forms.gle" in u or "google.com/forms" in u), "")
    spreadsheets = [u for u in urls if "docs.google.com/spreadsheets" in u]
    if not bundle.csv and spreadsheets:
        bundle.csv = spreadsheets[0]
    if not bundle.sheet and spreadsheets:
        bundle.sheet = spreadsheets[-1]
    return bundle


def first_url(text: str) -> str:
    match = URL_RE.search(text)
    return clean_url(match.group(0)) if match else ""


def clean_url(url: str) -> str:
    return html.unescape(url).rstrip(".,);")


def unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def evaluate_group(group: str, links: LinkBundle, feedback_text: str, dream_text: str) -> tuple[list[RubricItem], list[str], list[str]]:
    flags: list[str] = []
    notes: list[str] = []
    rubric: list[RubricItem] = []

    feedback_score, feedback_status, feedback_evidence, feedback_flags = score_feedback_correction(feedback_text, dream_text)
    rubric.append(RubricItem("1.1 Feedbacks do T1 corrigidos no DreamShaper", 0.25, feedback_score, feedback_status, feedback_evidence, feedback_flags))
    flags.extend(feedback_flags)

    planning_score, planning_status, planning_evidence, planning_flags = score_planning(dream_text)
    rubric.append(RubricItem("1.2 Planejamento do Projeto no DreamShaper", 0.30, planning_score, planning_status, planning_evidence, planning_flags))
    flags.extend(planning_flags)

    rubric.append(score_link_item("1.3 Link do Forms ou instrumento equivalente", 0.15, links.forms, "forms", flags))
    rubric.append(score_link_item("1.4 Link do CSV bruto", 0.15, links.csv, "csv", flags))
    rubric.append(score_link_item("1.5 Link da Google Planilha", 0.15, links.sheet, "sheet", flags))

    rubric.extend(score_sheet_block())
    rubric.extend(score_interpretation_block())
    flags.extend(["planilha_conteudo_nao_verificado", "revisao_manual_obrigatoria"])
    notes.append(
        "Pontuacao local cobre AVA, existencia dos links, feedback T1 e DreamShaper. "
        "Os blocos da Google Planilha e interpretacao permanecem pendentes ate acesso/download dos links."
    )
    if links.csv and links.sheet and spreadsheet_id(links.csv) == spreadsheet_id(links.sheet):
        flags.append("csv_e_planilha_mesmo_link")
        notes.append("O link informado como CSV e o link da planilha apontam para o mesmo arquivo; verificar se o CSV bruto foi preservado.")
    return rubric, flags, notes


def score_feedback_correction(feedback_text: str, dream_text: str) -> tuple[float, str, str, list[str]]:
    if not feedback_text:
        return 0.0, "bloqueado", "Feedback oficial do T1 nao encontrado.", ["feedback_t1_nao_verificavel"]
    if not dream_text:
        return 0.0, "bloqueado", "DreamShaper atual nao encontrado.", ["dreamshaper_nao_verificado"]

    text = dream_text.lower()
    evidence: list[str] = []
    points = 0.0
    if "pergunta de pesquisa" in text or "?" in dream_text:
        points += 0.05
        evidence.append("ha pergunta/problema de pesquisa explicitado")
    if any(term in text for term in ["público-alvo", "publico-alvo", "amostra", "participantes"]):
        points += 0.04
        evidence.append("ha publico-alvo/amostra descritos")
    if any(term in text for term in ["variáveis observadas", "variaveis observadas", "indicadores", "renda", "proficiência", "proficiencia"]):
        points += 0.05
        evidence.append("ha variaveis/indicadores mencionados")
    if any(term in text for term in ["questionário", "questionario", "forms", "formulário", "formulario"]):
        points += 0.04
        evidence.append("ha instrumento de coleta/metodologia")
    if planning_has_required_content(dream_text):
        points += 0.04
        evidence.append("ha algum preenchimento substantivo em Planejamento do Projeto")
    if ("tipo" in text or "escala" in text) and ("variável" in text or "variavel" in text):
        points += 0.04
        evidence.append("ha alguma indicacao de tipo/escala das variaveis")

    points = min(points, 0.25)
    flags: list[str] = []
    if points < 0.20:
        flags.append("feedback_t1_parcialmente_corrigido")
    status = "atendido" if points >= 0.23 else "parcial" if points >= 0.10 else "fragil"
    return points, status, "; ".join(evidence) if evidence else "Nao ha evidencias suficientes.", flags


def score_planning(dream_text: str) -> tuple[float, str, str, list[str]]:
    if not dream_text:
        return 0.0, "bloqueado", "DreamShaper atual nao encontrado.", ["dreamshaper_nao_verificado"]
    planning = extract_section(dream_text, "Planejamento do Projeto", ["Implementação do Projeto", "Implementacao do Projeto"])
    text = planning.lower()
    evidence = []
    points = 0.0
    checks = [
        (0.06, ["cronograma", "data", "prazo", "etapa"], "etapas/cronograma"),
        (0.06, ["recurso", "forms", "planilha", "questionário", "questionario"], "recursos/instrumentos"),
        (0.06, ["divisão", "divisao", "responsável", "responsavel", "equipe"], "divisao de tarefas"),
        (0.06, ["coleta", "export", "organiza", "analise", "análise"], "coleta/organizacao/analise"),
        (0.06, ["participantes", "amostra", "entrevista", "formulário", "formulario"], "procedimento de aplicacao"),
    ]
    for value, keywords, label in checks:
        if any(keyword in text for keyword in keywords):
            points += value
            evidence.append(label)
    if not planning.strip():
        return 0.0, "nao_atendido", "A secao Planejamento do Projeto aparece vazia no texto extraido.", ["planejamento_incompleto"]
    flags = [] if points >= 0.20 else ["planejamento_incompleto"]
    status = "atendido" if points >= 0.26 else "parcial" if points >= 0.10 else "fragil"
    return min(points, 0.30), status, "; ".join(evidence) if evidence else "Planejamento sem elementos suficientes.", flags


def planning_has_required_content(text: str) -> bool:
    planning = extract_section(text, "Planejamento do Projeto", ["Implementação do Projeto", "Implementacao do Projeto"])
    if len(planning) < 80:
        return False
    planning_lower = planning.lower()
    return any(
        keyword in planning_lower
        for keyword in ["cronograma", "etapa", "recurso", "divisão", "divisao", "coleta", "planilha", "forms"]
    )


def extract_section(text: str, heading: str, next_headings: list[str]) -> str:
    start = re.search(re.escape(heading), text, re.IGNORECASE)
    if not start:
        return ""
    section = text[start.end() :]
    end_positions = [m.start() for h in next_headings for m in [re.search(re.escape(h), section, re.IGNORECASE)] if m]
    if end_positions:
        section = section[: min(end_positions)]
    return strip_page_chrome(section)


def strip_page_chrome(text: str) -> str:
    lines = []
    for line in text.splitlines():
        clean = line.strip()
        if not clean:
            continue
        if clean.startswith("## Pagina") or clean.startswith("### Tabela"):
            continue
        if "UNIFOR: Atividade de Extensão" in clean:
            continue
        if re.fullmatch(r"20\d{2}\.1", clean):
            continue
        lines.append(clean)
    return "\n".join(lines).strip()


def score_link_item(item: str, max_score: float, url: str, kind: str, flags: list[str]) -> RubricItem:
    if not url:
        flags.append(f"{kind}_ausente")
        return RubricItem(item, max_score, 0.0, "ausente", "Link nao localizado no AVA.", [f"{kind}_ausente"])
    item_flags: list[str] = ["link_acesso_nao_verificado"]
    score = max_score
    evidence = f"Link localizado: {url}"
    status = "presente_nao_verificado"
    if kind == "forms" and not ("forms.gle" in url or "google.com/forms" in url):
        score = max_score / 2
        item_flags.append("link_tipo_incorreto")
        status = "parcial"
    if kind == "csv" and "docs.google.com/spreadsheets" in url:
        score = max_score / 2
        item_flags.append("csv_link_planilha")
        status = "parcial"
        evidence += " | O link rotulado como CSV aponta para Google Planilha; verificar se ha CSV bruto preservado."
    if kind == "sheet" and "docs.google.com/spreadsheets" not in url:
        score = max_score / 2
        item_flags.append("link_tipo_incorreto")
        status = "parcial"
    flags.extend(item_flags)
    return RubricItem(item, max_score, score, status, evidence, item_flags)


def score_sheet_block() -> list[RubricItem]:
    return [
        RubricItem("2.1 Organizacao, importacao e documentacao inicial", 0.35, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("2.2 Dicionario de variaveis", 0.40, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("2.3 Inspecao de qualidade dos dados", 0.45, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("2.4 Base preparada para analise", 0.30, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("2.5 Analise univariada qualitativa", 0.50, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("2.6 Analise univariada quantitativa", 0.45, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("2.7 Sintese exploratoria inicial", 0.35, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
    ]


def score_interpretation_block() -> list[RubricItem]:
    return [
        RubricItem("3.1 Interpretacao das distribuicoes univariadas", 0.40, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("3.2 Relacao das variaveis com a pergunta", 0.20, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("3.3 Clareza da conclusao descritiva", 0.15, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
        RubricItem("3.4 Limitacoes da coleta, amostra e dados", 0.25, 0.0, "pendente", "Depende de acesso a Google Planilha.", ["planilha_conteudo_nao_verificado"]),
    ]


def spreadsheet_id(url: str) -> str:
    match = SPREADSHEET_ID_RE.search(url or "")
    return match.group(1) if match else ""


def apply_remote_audit(audits: list[GroupAudit], out: Path) -> None:
    diagnostics: list[dict[str, Any]] = []
    for audit in audits:
        group_slug = audit.group.replace(" ", "_")
        group_dir = out / "downloads" / group_slug
        group_dir.mkdir(parents=True, exist_ok=True)
        diag: dict[str, Any] = {"group": audit.group, "links": {}, "sheet_audit": None}

        if audit.links.forms:
            form_diag = check_http_access(audit.links.forms)
            diag["links"]["forms"] = form_diag
            update_link_access(audit, "1.3 Link do Forms ou instrumento equivalente", form_diag)

        csv_diag: dict[str, Any] | None = None
        if audit.links.csv:
            csv_diag = download_spreadsheet_link(audit.links.csv, group_dir / "csv_link.xlsx")
            diag["links"]["csv"] = csv_diag
            update_link_access(audit, "1.4 Link do CSV bruto", csv_diag)

        sheet_diag: dict[str, Any] | None = None
        if audit.links.sheet:
            sheet_diag = download_spreadsheet_link(audit.links.sheet, group_dir / "google_planilha.xlsx")
            diag["links"]["sheet"] = sheet_diag
            update_link_access(audit, "1.5 Link da Google Planilha", sheet_diag)
            if sheet_diag.get("downloaded"):
                sheet_items, sheet_flags, sheet_notes = audit_workbook(Path(sheet_diag["path"]))
                replace_items(audit, sheet_items)
                audit.flags.extend(sheet_flags)
                audit.notes = [
                    note
                    for note in audit.notes
                    if "permanecem pendentes ate acesso/download dos links" not in note
                ]
                audit.notes.append(f"Google Planilha baixada e auditada em `{sheet_diag['path']}`.")
                audit.notes.extend(sheet_notes)
                stat_flags, stat_notes = audit_statistical_calculations(
                    audit.group,
                    Path(sheet_diag["path"]),
                    out / "auditoria_estatistica" / f"{audit.group.replace(' ', '_')}_calculos.md",
                )
                audit.flags.extend(stat_flags)
                audit.notes.extend(stat_notes)
                apply_statistical_penalties(audit, stat_flags)
                interp_items, interp_flags, interp_notes = audit_interpretations(
                    audit.group,
                    Path(sheet_diag["path"]),
                    out / "auditoria_interpretativa" / f"{audit.group.replace(' ', '_')}_interpretacao.md",
                )
                replace_items(audit, interp_items)
                audit.flags.extend(interp_flags)
                audit.notes.extend(interp_notes)
                diag["sheet_audit"] = {"flags": sheet_flags, "notes": sheet_notes}

        audit.flags = sorted(set(audit.flags))
        audit.flags = recompute_flags(audit)
        audit.rubric = round_rubric_items(audit.rubric)
        audit.local_score = sum(item.score for item in audit.rubric)
        audit.local_score_rounded = round_up_to_step(audit.local_score, 0.05)
        if not any(item.status == "pendente" for item in audit.rubric):
            audit.final_status = "preliminar_com_auditoria_remota"
        diagnostics.append(diag)

    (out / "links_diagnostico.json").write_text(json.dumps(diagnostics, ensure_ascii=False, indent=2), encoding="utf-8")


def check_http_access(url: str) -> dict[str, Any]:
    try:
        response = requests.get(url, timeout=20, allow_redirects=True)
        return {
            "url": url,
            "ok": response.ok,
            "status_code": response.status_code,
            "content_type": response.headers.get("content-type", ""),
            "final_url": response.url,
            "bytes": len(response.content),
        }
    except requests.RequestException as exc:
        return {"url": url, "ok": False, "error": exc.__class__.__name__, "message": str(exc)}


def download_spreadsheet_link(url: str, output: Path) -> dict[str, Any]:
    sid = spreadsheet_id(url)
    if not sid:
        diag = check_http_access(url)
        diag["downloaded"] = False
        diag["reason"] = "sem_id_google_spreadsheet"
        return diag
    export_url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"
    try:
        response = requests.get(export_url, timeout=40, allow_redirects=True)
        content = response.content
        is_xlsx = content.startswith(b"PK")
        diag = {
            "url": url,
            "export_url": export_url,
            "ok": response.ok,
            "status_code": response.status_code,
            "content_type": response.headers.get("content-type", ""),
            "bytes": len(content),
            "downloaded": bool(response.ok and is_xlsx),
            "path": str(output),
        }
        if response.ok and is_xlsx:
            output.write_bytes(content)
        else:
            diag["reason"] = "resposta_nao_xlsx_ou_sem_permissao"
            diag["preview"] = content[:300].decode("utf-8", errors="replace")
        return diag
    except requests.RequestException as exc:
        return {"url": url, "export_url": export_url, "ok": False, "downloaded": False, "error": exc.__class__.__name__, "message": str(exc)}


def update_link_access(audit: GroupAudit, item_name: str, diag: dict[str, Any]) -> None:
    for item in audit.rubric:
        if item.item == item_name:
            if diag.get("ok"):
                item.flags = [flag for flag in item.flags if flag != "link_acesso_nao_verificado"]
                item.evidence += f" | Acesso HTTP verificado ({diag.get('status_code')})."
                item.status = "verificado" if item.status == "presente_nao_verificado" else item.status
            else:
                reason = diag.get("message", "") or diag.get("reason", "") or str(diag.get("status_code", ""))
                if "Failed to resolve" in reason or diag.get("error") in {"ConnectionError", "Timeout"}:
                    item.flags.append("falha_automacao_link")
                    item.evidence += " | A automacao nao conseguiu verificar o link por falha de rede/ambiente; nao zerar sem revisao manual."
                    audit.flags.append("falha_automacao_link")
                else:
                    item.flags.append("link_sem_acesso")
                    item.status = "bloqueado"
                    item.score = 0.0
                    item.evidence += f" | Falha de acesso: {diag.get('error') or diag.get('status_code') or diag.get('reason')}."
                    audit.flags.append("link_sem_acesso")


def replace_items(audit: GroupAudit, replacements: list[RubricItem]) -> None:
    by_name = {item.item: item for item in replacements}
    audit.rubric = [by_name.get(item.item, item) for item in audit.rubric]


def round_rubric_items(items: list[RubricItem]) -> list[RubricItem]:
    for item in items:
        item.score = min(item.max_score, round_up_to_step(item.score, 0.05))
    return items


def recompute_flags(audit: GroupAudit) -> list[str]:
    special = {
        flag
        for flag in audit.flags
        if flag
        not in {
            "link_acesso_nao_verificado",
            "planilha_conteudo_nao_verificado",
            "falha_automacao_link",
        }
    }
    item_flags = {flag for item in audit.rubric for flag in item.flags}
    if not any(item.status == "pendente" for item in audit.rubric):
        item_flags.discard("planilha_conteudo_nao_verificado")
    if not any("link_acesso_nao_verificado" in item.flags for item in audit.rubric):
        item_flags.discard("link_acesso_nao_verificado")
    return sorted(special | item_flags)


def apply_statistical_penalties(audit: GroupAudit, stat_flags: list[str]) -> None:
    if "frequencias_inconsistentes" in stat_flags or "percentuais_inconsistentes" in stat_flags:
        cap_item_score(audit, "2.5 Analise univariada qualitativa", 0.30, "Pontuacao limitada por inconsistencias em frequencias/percentuais.")
    if "medidas_quantitativas_nao_conferem" in stat_flags:
        cap_item_score(audit, "2.6 Analise univariada quantitativa", 0.25, "Pontuacao limitada por divergencia nas medidas recalculadas.")
    if "n_total_inconsistente" in stat_flags:
        cap_item_score(audit, "2.3 Inspecao de qualidade dos dados", 0.25, "Pontuacao limitada por inconsistencia no N total declarado.")
    if "base_preparada_tamanho_inconsistente" in stat_flags:
        cap_item_score(audit, "2.3 Inspecao de qualidade dos dados", 0.25, "Pontuacao limitada por inconsistencia de tamanho amostral.")
        cap_item_score(audit, "2.4 Base preparada para analise", 0.20, "Pontuacao limitada por inconsistencia entre base bruta/preparada.")


def cap_item_score(audit: GroupAudit, item_name: str, cap: float, reason: str) -> None:
    for item in audit.rubric:
        if item.item == item_name and item.score > cap:
            item.score = cap
            item.flags.append("auditoria_estatistica_ajuste")
            item.evidence = f"{item.evidence}; {reason}" if item.evidence else reason


def audit_workbook(path: Path) -> tuple[list[RubricItem], list[str], list[str]]:
    flags: list[str] = []
    notes: list[str] = []
    xls = pd.ExcelFile(path)
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet_names = xls.sheet_names
    normalized = canonical_sheet_map(sheet_names)
    required = [
        "01_csv_bruto",
        "02_dicionario_variaveis",
        "03_inspecao_qualidade",
        "04_dados_preparados",
        "05_qualitativas",
        "06_quantitativas",
        "07_sintese_univariada",
    ]
    missing = [name for name in required if name not in normalized]
    if missing:
        flags.append("aba_obrigatoria_ausente")
        notes.append(f"Abas obrigatorias ausentes: {', '.join(missing)}.")
    notes.append(
        "Mapeamento de abas: "
        + "; ".join(f"{canonical} -> {real}" for canonical, real in sorted(normalized.items()))
    )

    sheet_texts = {key: worksheet_text(workbook[real_name]) for key, real_name in normalized.items()}
    interpretation_text = "\n".join(
        sheet_texts.get(name, "")
        for name in ["05_qualitativas", "06_quantitativas", "07_sintese_univariada", "08_conclusao"]
    ).lower()
    items = [
        score_organization(normalized, required, sheet_texts, missing),
        score_dictionary(sheet_texts),
        score_quality(sheet_texts),
        score_prepared_base(sheet_texts, normalized),
        score_qualitative(sheet_texts, workbook, normalized),
        score_quantitative(sheet_texts, workbook, normalized),
        score_synthesis(sheet_texts),
        score_interp_distributions(interpretation_text),
        score_relation_question(interpretation_text),
        score_conclusion(interpretation_text),
        score_limitations(interpretation_text),
    ]
    for item in items:
        flags.extend(item.flags)
    return items, sorted(set(flags)), notes


def audit_statistical_calculations(group: str, path: Path, report_path: Path) -> tuple[list[str], list[str]]:
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    canonical = canonical_sheet_map(wb_values.sheetnames)
    flags: list[str] = []
    notes: list[str] = []
    lines: list[str] = [f"# Auditoria Estatistica - {group}", "", f"Arquivo: `{path}`", ""]

    dataframes = {key: dataframe_from_worksheet(wb_values[real]) for key, real in canonical.items()}
    raw_n = dataframe_valid_rows(dataframes.get("01_csv_bruto"))
    prepared_n = dataframe_valid_rows(dataframes.get("04_dados_preparados"))
    lines.extend(["## Tamanho Das Bases", "", f"- `01_csv_bruto`: {raw_n} linhas validas", f"- `04_dados_preparados`: {prepared_n} linhas validas", ""])
    if raw_n and prepared_n and abs(raw_n - prepared_n) > max(2, raw_n * 0.05):
        flags.append("base_preparada_tamanho_inconsistente")
        lines.append(f"- **Alerta:** diferenca relevante entre base bruta ({raw_n}) e preparada ({prepared_n}).")
        notes.append(f"Auditoria estatistica: base bruta tem {raw_n} linhas validas e base preparada tem {prepared_n}.")

    quality_text = worksheet_text(wb_values[canonical["03_inspecao_qualidade"]]) if "03_inspecao_qualidade" in canonical else ""
    quality_numbers = [float(match) for match in re.findall(r"\b\d+(?:[.,]\d+)?\b", quality_text.replace(",", "."))]
    likely_n_values = [n for n in quality_numbers if n >= 20]
    if prepared_n and likely_n_values and not any(abs(n - prepared_n) <= 2 for n in likely_n_values):
        flags.append("n_total_inconsistente")
        common_n = most_common_number(likely_n_values)
        lines.append(
            f"- **Alerta:** a base preparada tem {prepared_n} observacoes, mas a inspecao de qualidade declara principalmente N={common_n:g}. "
            f"Nao encontrei, nessa aba, um N compativel com {prepared_n}."
        )
        notes.append(
            f"Auditoria estatistica: dataset/base preparada tem {prepared_n} observacoes, mas a inspecao de qualidade declara principalmente N={common_n:g}; por isso o item 2.3 fica parcial."
        )

    formula_summary = summarize_formulas(wb_formulas, canonical)
    lines.extend(["", "## Formulas Detectadas", ""])
    for sheet, summary in formula_summary.items():
        lines.append(f"- `{sheet}`: {summary['total']} formulas; funcoes: {', '.join(summary['functions']) if summary['functions'] else 'nenhuma funcao detectada'}")
    lines.append("")

    qual_result = audit_qualitative_tables(wb_values, wb_formulas, canonical)
    lines.extend(["## Frequencias E Percentuais Qualitativos", ""])
    lines.extend(qual_result["lines"])
    flags.extend(qual_result["flags"])

    quant_result = audit_quantitative_measures(dataframes.get("04_dados_preparados"), wb_values, wb_formulas, canonical)
    lines.extend(["", "## Medidas Quantitativas", ""])
    lines.extend(quant_result["lines"])
    flags.extend(quant_result["flags"])

    if not flags:
        lines.extend(["", "## Resultado", "", "Nao foram encontradas divergencias computacionais relevantes nas checagens automatizadas."])
    else:
        lines.extend(["", "## Flags", ""])
        for flag in sorted(set(flags)):
            lines.append(f"- `{flag}`")

    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    if flags:
        notes.append(f"Auditoria estatistica gerou flags: {', '.join(sorted(set(flags)))}.")
    else:
        notes.append("Auditoria estatistica automatica nao encontrou divergencias relevantes.")
    return sorted(set(flags)), notes


def audit_interpretations(group: str, path: Path, report_path: Path) -> tuple[list[RubricItem], list[str], list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    canonical = canonical_sheet_map(wb.sheetnames)
    sections = extract_interpretive_sections(wb, canonical)
    combined = "\n\n".join(section["text"] for section in sections).strip()
    lower = combined.lower()
    flags: list[str] = []
    notes: list[str] = []

    item31 = score_interpretive_dimension(
        "3.1 Interpretacao das distribuicoes univariadas",
        0.40,
        lower,
        [
            ("leitura_resultados", ["maior", "menor", "predomin", "distribui", "observa", "frequ"]),
            ("quantifica", ["%", "percent", "respostas", "n=", "válidas", "validas"]),
            ("amostra", ["amostra", "respondentes", "participantes"]),
            ("sem_extrapolacao", ["amostra", "respondentes", "dados coletados", "respostas coletadas"]),
        ],
        "interpretacao_fragil",
    )
    item32 = score_interpretive_dimension(
        "3.2 Relacao das variaveis com a pergunta",
        0.20,
        lower,
        [
            ("pergunta_objetivo", ["pergunta", "objetivo", "pesquisa"]),
            ("variaveis", ["variável", "variavel", "curso", "renda", "pmg", "internacionalização", "internacionalizacao"]),
            ("instrumento_dados", ["questionário", "questionario", "forms", "dados", "coleta"]),
            ("alinhamento", ["relação", "relacao", "coerente", "indica", "permite observar"]),
        ],
        "pergunta_variaveis_desalinhadas",
    )
    item33 = score_interpretive_dimension(
        "3.3 Clareza da conclusao descritiva",
        0.15,
        lower,
        [
            ("conclusao", ["conclus", "síntese", "sintese"]),
            ("descritiva", ["descrit", "observ", "dados", "amostra"]),
            ("proximos_passos", ["próximo", "proximo", "bivariada", "futura"]),
        ],
        "conclusao_fragil",
    )
    item34 = score_interpretive_dimension(
        "3.4 Limitacoes da coleta, amostra e dados",
        0.25,
        lower,
        [
            ("limitacoes", ["limita", "restri", "cautela"]),
            ("amostra", ["amostra", "respondentes", "participantes"]),
            ("ausentes_inconsistencias", ["ausente", "nulo", "inconsist", "inválid", "invalid"]),
            ("nao_causalidade", ["não permite concluir causa", "nao permite concluir causa", "não causal", "nao causal", "associação não implica", "associacao nao implica"]),
            ("generalizacao", ["não generaliza", "nao generaliza", "dados coletados", "amostra coletada"]),
        ],
        "limitacoes_fragil",
    )

    causal_terms = ["influencia diretamente", "impacta o acesso", "decide quem", "causa", "efeito"]
    if any(term in lower for term in causal_terms) and not any(term in lower for term in ["não causal", "nao causal", "não permite concluir causa", "nao permite concluir causa"]):
        item34.flags.append("conclusao_causal_revisao")
        flags.append("conclusao_causal_revisao")

    items = [item31, item32, item33, item34]
    for item in items:
        flags.extend(item.flags)

    lines = [f"# Auditoria Interpretativa - {group}", "", f"Arquivo: `{path}`", ""]
    lines.extend(["## Secoes Textuais Encontradas", ""])
    if sections:
        for section in sections:
            excerpt = section["text"][:700].replace("\n", " ")
            lines.extend([f"### {section['sheet']}:{section['cell']}", "", excerpt, ""])
    else:
        lines.append("Nenhuma secao textual interpretativa suficiente foi localizada nas abas analiticas.")
    lines.extend(["", "## Pontuacao Interpretativa", ""])
    for item in items:
        lines.append(f"- {item.item}: {item.score:.2f} / {item.max_score:.2f} ({item.evidence})")
    if flags:
        lines.extend(["", "## Flags", ""])
        for flag in sorted(set(flags)):
            lines.append(f"- `{flag}`")
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    notes.append(f"Auditoria interpretativa registrada em `{report_path}`.")
    return items, sorted(set(flags)), notes


def extract_interpretive_sections(wb: Any, canonical: dict[str, str]) -> list[dict[str, str]]:
    sections: list[dict[str, str]] = []
    target_sheets = ["05_qualitativas", "06_quantitativas", "07_sintese_univariada", "08_conclusao"]
    keywords = ["interpre", "observa", "conclus", "síntese", "sintese", "limita", "próximo", "proximo"]
    for canonical_name in target_sheets:
        real_name = canonical.get(canonical_name)
        if not real_name:
            continue
        ws = wb[real_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                text = value.strip()
                lower = text.lower()
                if len(text) >= 45 and (any(keyword in lower for keyword in keywords) or looks_like_interpretive_sentence(text)):
                    sections.append({"sheet": real_name, "cell": cell.coordinate, "text": text})
    return sections


def looks_like_interpretive_sentence(text: str) -> bool:
    lower = text.lower()
    return len(text.split()) >= 10 and any(term in lower for term in ["dados", "respostas", "amostra", "resultado", "grupo", "particip"])


def score_interpretive_dimension(
    item_name: str,
    max_score: float,
    text: str,
    dimensions: list[tuple[str, list[str]]],
    weak_flag: str,
) -> RubricItem:
    found = [label for label, terms in dimensions if any(term in text for term in terms)]
    score = max_score * len(found) / len(dimensions)
    flags = [] if score >= max_score * 0.6 else [weak_flag]
    evidence = "Dimensoes/evidencias: " + ", ".join(found) if found else "Nao foram localizadas evidencias interpretativas suficientes."
    return RubricItem(item_name, max_score, min(score, max_score), "verificado", evidence, flags)


def most_common_number(values: list[float]) -> float:
    counts: dict[float, int] = {}
    for value in values:
        rounded = round(value, 6)
        counts[rounded] = counts.get(rounded, 0) + 1
    return max(counts, key=counts.get)


def dataframe_from_worksheet(ws: Any) -> pd.DataFrame:
    rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    header_idx = find_header_row(rows)
    if header_idx is None:
        return pd.DataFrame()
    header = [normalize_header(value, idx) for idx, value in enumerate(rows[header_idx])]
    data = rows[header_idx + 1 :]
    df = pd.DataFrame(data, columns=header)
    df = df.dropna(how="all")
    df = df.loc[:, [not str(col).startswith("vazio_") or df[col].notna().any() for col in df.columns]]
    return df


def find_header_row(rows: list[list[Any]]) -> int | None:
    best_idx = None
    best_score = 0
    for idx, row in enumerate(rows[:15]):
        nonempty = sum(1 for value in row if value not in (None, ""))
        text = " ".join(str(value).lower() for value in row if value not in (None, ""))
        score = nonempty
        if any(term in text for term in ["carimbo", "vari", "categoria", "freq", "id", "data"]):
            score += 5
        if score > best_score and nonempty >= 2:
            best_idx = idx
            best_score = score
    return best_idx


def normalize_header(value: Any, idx: int) -> str:
    if value in (None, ""):
        return f"vazio_{idx+1}"
    text = str(value).strip()
    return text or f"vazio_{idx+1}"


def dataframe_valid_rows(df: pd.DataFrame | None) -> int:
    if df is None or df.empty:
        return 0
    meaningful_cols = [col for col in df.columns if not str(col).startswith("vazio_")]
    if meaningful_cols:
        first_cols = meaningful_cols[: min(3, len(meaningful_cols))]
        return int(df[first_cols].dropna(how="all").shape[0])
    return int(df.dropna(how="all").shape[0])


def summarize_formulas(wb: Any, canonical: dict[str, str]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    formula_re = re.compile(r"=([A-Z][A-Z0-9.]*)\(")
    for canonical_name in ["03_inspecao_qualidade", "05_qualitativas", "06_quantitativas", "07_sintese_univariada"]:
        real_name = canonical.get(canonical_name)
        if not real_name:
            continue
        ws = wb[real_name]
        formulas = []
        functions = set()
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append(value)
                    functions.update(match.group(1).upper() for match in formula_re.finditer(value))
        result[canonical_name] = {"total": len(formulas), "functions": sorted(functions)}
    return result


def audit_qualitative_tables(wb_values: Any, wb_formulas: Any, canonical: dict[str, str]) -> dict[str, Any]:
    lines: list[str] = []
    flags: list[str] = []
    real_name = canonical.get("05_qualitativas")
    if not real_name:
        return {"lines": ["- Aba qualitativa nao localizada."], "flags": ["sem_analise_univariada"]}
    ws = wb_values[real_name]
    wsf = wb_formulas[real_name]
    formula_count = count_formulas(wsf)
    table_checks = scan_frequency_tables(ws)
    lines.append(f"- Aba auditada: `{real_name}`.")
    lines.append(f"- Formulas na aba: {formula_count}.")
    if not table_checks:
        flags.append("frequencias_nao_verificaveis")
        lines.append("- Nao foram encontradas tabelas com cabecalhos Categoria/Freq/% verificaveis.")
    ok_tables = 0
    divergent_tables = 0
    for check in table_checks:
        lines.append(
            f"- Tabela em {check['origin']}: soma frequencias={check['freq_sum']}, total={check['total_freq']}, soma percentuais={check['pct_sum']:.4f}, status={check['status']}."
        )
        if check["freq_ok"] and check["pct_ok"]:
            ok_tables += 1
        else:
            divergent_tables += 1
    if table_checks and ok_tables == 0:
        flags.append("frequencias_inconsistentes")
    if table_checks and divergent_tables > ok_tables:
        flags.append("percentuais_inconsistentes")
    if formula_count == 0:
        flags.append("formulas_qualitativas_ausentes")
    return {"lines": lines, "flags": sorted(set(flags))}


def count_formulas(ws: Any) -> int:
    return sum(1 for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("="))


def scan_frequency_tables(ws: Any) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    for row in range(1, ws.max_row + 1):
        values = [normalize_sheet_name(str(ws.cell(row, col).value or "")) for col in range(1, ws.max_column + 1)]
        for col in range(1, ws.max_column - 1):
            trio = values[col - 1 : col + 2]
            if "categoria" in trio[0] and ("freq" in trio[1] or "frequencia" in trio[1]) and trio[2] in {"", "%", "percentual"}:
                check = check_frequency_table(ws, row, col, col + 1, col + 2)
                if check:
                    checks.append(check)
            if values[col - 1] and any("freq" in value or "frequencia" in value for value in values[col : col + 4]):
                freq_col = None
                pct_col = None
                for offset in range(1, 5):
                    candidate_col = col + offset
                    if candidate_col > ws.max_column:
                        continue
                    label = values[candidate_col - 1]
                    if freq_col is None and ("freq" in label or "frequencia" in label):
                        freq_col = candidate_col
                    if pct_col is None and (label == "" or "percent" in label or label == "%"):
                        header_raw = str(ws.cell(row, candidate_col).value or "").strip()
                        if header_raw in {"", "%"} or "percent" in normalize_sheet_name(header_raw):
                            pct_col = candidate_col
                if freq_col and pct_col and freq_col != pct_col:
                    check = check_frequency_table(ws, row, col, freq_col, pct_col)
                    if check:
                        checks.append(check)
    unique: dict[str, dict[str, Any]] = {}
    for check in checks:
        unique[check["origin"]] = check
    return list(unique.values())


def check_frequency_table(ws: Any, header_row: int, category_col: int, freq_col: int, pct_col: int) -> dict[str, Any] | None:
    freq_values: list[float] = []
    pct_values: list[float] = []
    total_freq = None
    total_pct = None
    for row in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
        category = ws.cell(row, category_col).value
        freq = as_number(ws.cell(row, freq_col).value)
        pct = as_number(ws.cell(row, pct_col).value)
        if category in (None, "") and freq is None and pct is None:
            if freq_values:
                break
            continue
        category_text = str(category or "").strip().lower()
        if "interpreta" in category_text or "observa" in category_text:
            break
        if "total" in category_text:
            total_freq = freq
            total_pct = pct
            break
        if category_text and freq is not None:
            freq_values.append(freq)
            if pct is not None:
                pct_values.append(pct)
    if not freq_values:
        return None
    freq_sum = sum(freq_values)
    pct_sum = sum(pct_values)
    freq_ok = total_freq is None or abs(freq_sum - total_freq) <= 1e-6
    pct_ok = not pct_values or total_pct is None or abs(pct_sum - total_pct) <= 0.02
    return {
        "origin": ws.cell(header_row, category_col).coordinate,
        "freq_sum": freq_sum,
        "total_freq": total_freq,
        "pct_sum": pct_sum,
        "total_pct": total_pct,
        "freq_ok": freq_ok,
        "pct_ok": pct_ok,
        "status": "ok" if freq_ok and pct_ok else "divergente",
    }


def audit_quantitative_measures(prepared_df: pd.DataFrame | None, wb_values: Any, wb_formulas: Any, canonical: dict[str, str]) -> dict[str, Any]:
    lines: list[str] = []
    flags: list[str] = []
    real_name = canonical.get("06_quantitativas")
    if not real_name:
        return {"lines": ["- Aba quantitativa nao localizada."], "flags": ["aba_obrigatoria_ausente"]}
    ws = wb_values[real_name]
    wsf = wb_formulas[real_name]
    formula_count = count_formulas(wsf)
    formulas = collect_formula_texts(wsf)
    lines.append(f"- Aba auditada: `{real_name}`.")
    lines.append(f"- Formulas na aba: {formula_count}.")
    if formula_count == 0:
        flags.append("formulas_quantitativas_ausentes")

    displayed_numbers = collect_numbers(ws)
    numeric_stats = compute_numeric_stats(prepared_df)
    if not numeric_stats:
        lines.append("- Nao foi possivel identificar variaveis numericas na base preparada.")
        flags.append("medidas_quantitativas_nao_verificaveis")
    else:
        matched_any = False
        for col, stats in numeric_stats.items():
            matches = [name for name, value in stats.items() if number_present(value, displayed_numbers)]
            if matches:
                matched_any = True
            lines.append(f"- `{col}` recalculada: " + ", ".join(f"{k}={v:.4f}" for k, v in stats.items()) + f"; medidas encontradas na aba: {', '.join(matches) if matches else 'nenhuma'}")
        if not matched_any:
            flags.append("medidas_quantitativas_nao_conferem")

    expected_functions = ["AVERAGE", "MEDIAN", "MIN", "MAX", "STDEV"]
    used_functions = {fn for formula in formulas for fn in expected_functions if fn in formula.upper()}
    if numeric_stats and len(used_functions) < 3:
        flags.append("formulas_quantitativas_incompletas")
        lines.append(f"- Funcoes estatisticas nativas detectadas: {', '.join(sorted(used_functions)) if used_functions else 'nenhuma'}.")
    return {"lines": lines, "flags": sorted(set(flags))}


def collect_formula_texts(ws: Any) -> list[str]:
    return [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]


def collect_numbers(ws: Any) -> list[float]:
    numbers: list[float] = []
    for row in ws.iter_rows(values_only=True):
        for value in row:
            number = as_number(value)
            if number is not None:
                numbers.append(number)
    return numbers


def compute_numeric_stats(df: pd.DataFrame | None) -> dict[str, dict[str, float]]:
    if df is None or df.empty:
        return {}
    result: dict[str, dict[str, float]] = {}
    for col in df.columns:
        series = pd.to_numeric(df[col], errors="coerce").dropna()
        if len(series) < 5:
            continue
        unique_ratio = series.nunique() / max(len(series), 1)
        if unique_ratio < 0.02 and len(series) > 50:
            continue
        result[str(col)] = {
            "n": float(len(series)),
            "media": float(series.mean()),
            "mediana": float(series.median()),
            "min": float(series.min()),
            "max": float(series.max()),
            "desvio": float(series.std(ddof=1)) if len(series) > 1 else 0.0,
        }
    return result


def number_present(value: float, numbers: list[float], rel_tol: float = 0.02, abs_tol: float = 0.05) -> bool:
    return any(abs(candidate - value) <= max(abs_tol, abs(value) * rel_tol) for candidate in numbers)


def as_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) if isinstance(value, float) else False:
            return None
        return float(value)
    text = str(value).strip().replace("%", "")
    text = text.replace(".", "").replace(",", ".") if re.search(r"\d,\d", text) else text
    try:
        return float(text)
    except ValueError:
        return None


def normalize_sheet_name(name: str) -> str:
    text = unicodedata.normalize("NFKD", name.strip().lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def canonical_sheet_map(sheet_names: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for name in sheet_names:
        normalized = normalize_sheet_name(name)
        canonical = canonical_sheet_name(normalized)
        if canonical and canonical not in result:
            result[canonical] = name
    return result


def canonical_sheet_name(normalized: str) -> str:
    if re.search(r"(^|_)0?1(_|$)", normalized) and "csv" in normalized:
        return "01_csv_bruto"
    if re.search(r"(^|_)0?2(_|$)", normalized) and ("dicionario" in normalized or "variave" in normalized):
        return "02_dicionario_variaveis"
    if re.search(r"(^|_)0?3(_|$)", normalized) and ("inspecao" in normalized or "qualidade" in normalized):
        return "03_inspecao_qualidade"
    if re.search(r"(^|_)0?4(_|$)", normalized) and ("prepar" in normalized or "dados" in normalized):
        return "04_dados_preparados"
    if re.search(r"(^|_)0?5(_|$)", normalized) and "qualitativ" in normalized:
        return "05_qualitativas"
    if re.search(r"(^|_)0?6(_|$)", normalized) and "quantitativ" in normalized:
        return "06_quantitativas"
    if re.search(r"(^|_)0?7(_|$)", normalized) and ("sintese" in normalized or "univariada" in normalized):
        return "07_sintese_univariada"
    if re.search(r"(^|_)0?8(_|$)", normalized) and "conclus" in normalized:
        return "08_conclusao"

    if "csv" in normalized and "bruto" in normalized:
        return "01_csv_bruto"
    if "dicionario" in normalized and "variave" in normalized:
        return "02_dicionario_variaveis"
    if "inspecao" in normalized and "qualidade" in normalized:
        return "03_inspecao_qualidade"
    if "dados" in normalized and "prepar" in normalized:
        return "04_dados_preparados"
    if "qualitativ" in normalized:
        return "05_qualitativas"
    if "quantitativ" in normalized:
        return "06_quantitativas"
    if "sintese" in normalized and "univariad" in normalized:
        return "07_sintese_univariada"
    if "conclus" in normalized:
        return "08_conclusao"
    return ""


def worksheet_text(ws: Any) -> str:
    values: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if value is not None:
                values.append(str(value))
    return "\n".join(values)


def score_organization(normalized: dict[str, str], required: list[str], texts: dict[str, str], missing: list[str]) -> RubricItem:
    score = 0.0
    evidence = []
    flags = []
    present_count = len([name for name in required if name in normalized])
    score += 0.20 * present_count / len(required)
    evidence.append(f"{present_count}/{len(required)} abas obrigatorias localizadas")
    if "01_csv_bruto" in texts and count_nonempty_cells_text(texts["01_csv_bruto"]) >= 10:
        score += 0.10
        evidence.append("aba 01_csv_bruto contem dados")
    if present_count == len(required):
        score += 0.05
    if missing:
        flags.append("aba_obrigatoria_ausente")
    return RubricItem("2.1 Organizacao, importacao e documentacao inicial", 0.35, min(score, 0.35), "verificado", "; ".join(evidence), flags)


def count_nonempty_cells_text(text: str) -> int:
    return len([part for part in re.split(r"\s+", text) if part.strip()])


def score_dictionary(texts: dict[str, str]) -> RubricItem:
    text = texts.get("02_dicionario_variaveis", "").lower()
    terms = ["vari", "pergunta", "tipo", "escala", "valores", "papel"]
    score = sum(0.40 / len(terms) for term in terms if term in text)
    flags = [] if score >= 0.28 else ["sem_dicionario_variaveis"]
    return RubricItem("2.2 Dicionario de variaveis", 0.40, min(score, 0.40), "verificado", matched_terms_evidence(text, terms), flags)


def score_quality(texts: dict[str, str]) -> RubricItem:
    text = texts.get("03_inspecao_qualidade", "").lower()
    dimensions = [
        ("estrutura", ["linha", "coluna", "n total"]),
        ("validas", ["valid", "preench"]),
        ("ausentes", ["ausente", "nulo"]),
        ("duplicidades", ["duplic"]),
        ("inconsistencias", ["inconsist", "inválid", "invalid"]),
        ("amostra_observacoes", ["amost", "observa"]),
    ]
    found = [label for label, terms in dimensions if any(term in text for term in terms)]
    score = 0.45 * len(found) / len(dimensions)
    flags = [] if score >= 0.25 else ["sem_inspecao_qualidade"]
    evidence = "Dimensoes/evidencias: " + ", ".join(found) if found else "Nao foram localizadas evidencias suficientes."
    return RubricItem("2.3 Inspecao de qualidade dos dados", 0.45, min(score, 0.45), "verificado", evidence, flags)


def score_prepared_base(texts: dict[str, str], normalized: dict[str, str]) -> RubricItem:
    text = texts.get("04_dados_preparados", "").lower()
    score = 0.0
    evidence = []
    flags = []
    if "04_dados_preparados" in normalized:
        score += 0.15
        evidence.append("aba 04_dados_preparados localizada")
    if any(term in text for term in ["recod", "padron", "limpeza", "prepar", "ajuste"]):
        score += 0.10
        evidence.append("ha indicio de preparacao/decisoes")
    if "01_csv_bruto" in normalized:
        score += 0.05
        evidence.append("aba bruta preservada")
    if score < 0.15:
        flags.append("base_preparada_nao_verificada")
    return RubricItem("2.4 Base preparada para analise", 0.30, min(score, 0.30), "verificado", "; ".join(evidence), flags)


def score_qualitative(texts: dict[str, str], workbook: Any, normalized: dict[str, str]) -> RubricItem:
    text = texts.get("05_qualitativas", "").lower()
    score = 0.0
    evidence = []
    if "05_qualitativas" in normalized:
        score += 0.15
        evidence.append("aba 05_qualitativas localizada")
    if any(term in text for term in ["frequ", "contagem", "absolut"]):
        score += 0.12
        evidence.append("frequencia absoluta")
    if any(term in text for term in ["percent", "%", "relativ"]):
        score += 0.10
        evidence.append("frequencia relativa/percentual")
    if sheet_has_charts(workbook, normalized.get("05_qualitativas", "")):
        score += 0.06
        evidence.append("grafico detectado")
    if any(term in text for term in ["interpre", "observa", "distribui", "maior", "menor"]):
        score += 0.07
        evidence.append("interpretacao textual")
    flags = [] if score >= 0.30 else ["sem_analise_univariada"]
    return RubricItem("2.5 Analise univariada qualitativa", 0.50, min(score, 0.50), "verificado", "; ".join(evidence), flags)


def score_quantitative(texts: dict[str, str], workbook: Any, normalized: dict[str, str]) -> RubricItem:
    text = texts.get("06_quantitativas", "").lower()
    if "06_quantitativas" not in normalized:
        return RubricItem("2.6 Analise univariada quantitativa", 0.45, 0.0, "nao_verificado", "Aba 06_quantitativas ausente.", ["aba_obrigatoria_ausente"])
    if ("não há" in text or "nao ha" in text) and "quantitativa" in text:
        return RubricItem("2.6 Analise univariada quantitativa", 0.45, 0.0, "redistribuicao_manual", "Ausencia de quantitativas declarada; aplicar redistribuicao manual se justificativa for adequada.", ["quantitativas_ausentes_redistribuicao_manual"])
    terms = ["valid", "média", "media", "mediana", "mínimo", "minimo", "máximo", "maximo", "desvio"]
    score = 0.10 if "06_quantitativas" in normalized else 0.0
    score += min(0.28, sum(0.04 for term in terms if term in text))
    evidence = matched_terms_evidence(text, terms)
    if sheet_has_charts(workbook, normalized.get("06_quantitativas", "")):
        score += 0.04
        evidence += "; grafico detectado"
    if any(term in text for term in ["interpre", "observa", "distribui"]):
        score += 0.03
        evidence += "; interpretacao textual"
    flags = [] if score >= 0.25 else ["quantitativas_ausentes_sem_justificativa"]
    return RubricItem("2.6 Analise univariada quantitativa", 0.45, min(score, 0.45), "verificado", evidence, flags)


def score_synthesis(texts: dict[str, str]) -> RubricItem:
    text = texts.get("07_sintese_univariada", "").lower()
    terms = ["resultado", "pergunta", "limita", "amostra", "coleta", "bivariada", "próximo", "proximo"]
    score = sum(0.35 / 6 for term in terms if term in text)
    return RubricItem("2.7 Sintese exploratoria inicial", 0.35, min(score, 0.35), "verificado", matched_terms_evidence(text, terms), [] if score >= 0.20 else ["sintese_fragil"])


def score_interp_distributions(text: str) -> RubricItem:
    terms = ["interpre", "distribui", "respostas válidas", "respostas validas", "amostra", "respondentes"]
    score = min(0.40, sum(0.08 for term in terms if term in text))
    return RubricItem("3.1 Interpretacao das distribuicoes univariadas", 0.40, score, "verificado", matched_terms_evidence(text, terms), [] if score >= 0.24 else ["interpretacao_fragil"])


def score_relation_question(text: str) -> RubricItem:
    terms = ["pergunta", "objetivo", "variável", "variavel", "instrumento"]
    score = min(0.20, sum(0.04 for term in terms if term in text))
    return RubricItem("3.2 Relacao das variaveis com a pergunta", 0.20, score, "verificado", matched_terms_evidence(text, terms), [] if score >= 0.12 else ["pergunta_variaveis_desalinhadas"])


def score_conclusion(text: str) -> RubricItem:
    terms = ["conclus", "síntese", "sintese", "próximo", "proximo", "bivariada"]
    score = min(0.15, sum(0.03 for term in terms if term in text))
    return RubricItem("3.3 Clareza da conclusao descritiva", 0.15, score, "verificado", matched_terms_evidence(text, terms), [] if score >= 0.09 else ["conclusao_fragil"])


def score_limitations(text: str) -> RubricItem:
    terms = ["limita", "amostra", "ausente", "inconsist", "cautela", "causal", "generaliza"]
    score = min(0.25, sum(0.04 for term in terms if term in text))
    flags = [] if score >= 0.12 else ["limitacoes_fragil"]
    if any(term in text for term in ["causa", "influencia diretamente", "impacta o acesso"]):
        flags.append("conclusao_causal_revisao")
    return RubricItem("3.4 Limitacoes da coleta, amostra e dados", 0.25, score, "verificado", matched_terms_evidence(text, terms), flags)


def sheet_has_charts(workbook: Any, sheet_name: str) -> bool:
    if not sheet_name or sheet_name not in workbook.sheetnames:
        return False
    return bool(getattr(workbook[sheet_name], "_charts", []))


def matched_terms_evidence(text: str, terms: list[str]) -> str:
    found = [term for term in terms if term in text]
    return "Termos/evidencias: " + ", ".join(found) if found else "Nao foram localizadas evidencias suficientes."


def round_up_to_step(value: float, step: float) -> float:
    if value <= 0:
        return 0.0
    return math.ceil((value - 1e-9) / step) * step


def find_original_pdf(base: Path, dream_md: Path) -> str:
    if not dream_md:
        return ""
    stem = dream_md.stem
    candidates = sorted((base / "entradas" / "dreamshaper").glob(f"{stem}.pdf"))
    return str(candidates[0]) if candidates else ""


def write_pipeline_outputs(audits: list[GroupAudit], out: Path) -> None:
    (out / "manifest.json").write_text(
        json.dumps([audit_to_json(audit) for audit in audits], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_notes_csv(audits, out / "notas.csv")
    write_feedback_csv(audits, out / "feedback_moodle.csv")
    for audit in audits:
        write_group_rubric(audit, out / "rubricas" / f"rubrica_{audit.group.replace(' ', '_')}.md")
        write_group_feedback_html(audit, out / f"feedback_moodle_{audit.group.replace(' ', '_')}.html")
        write_group_feedback_md(audit, out / f"feedback_moodle_{audit.group.replace(' ', '_')}.md")
    write_summary(audits, out / "resumo_correcao.md")


def audit_to_json(audit: GroupAudit) -> dict[str, Any]:
    data = asdict(audit)
    return data


def write_notes_csv(audits: list[GroupAudit], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["grupo", "nota_local_preliminar", "nota_local_arredondada", "status", "flags"])
        for audit in audits:
            writer.writerow([audit.group, f"{audit.local_score:.3f}", f"{audit.local_score_rounded:.2f}", audit.final_status, "; ".join(audit.flags)])


def write_feedback_csv(audits: list[GroupAudit], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["grupo", "feedback_html"])
        for audit in audits:
            writer.writerow([audit.group, build_feedback_html_body(audit)])


def write_group_rubric(audit: GroupAudit, path: Path) -> None:
    lines = [
        f"# Rubrica Preliminar - {audit.group}",
        "",
        f"- Status: `{audit.final_status}`",
        f"- Nota local preliminar: `{audit.local_score:.3f} / 5,00`",
        f"- Nota local arredondada: `{audit.local_score_rounded:.2f} / 5,00`",
        f"- DreamShaper: `{audit.dreamshaper_md}`",
        "",
        "## Itens",
        "",
        "| Item | Maximo | Nota | Status | Evidencia | Flags |",
        "| --- | ---: | ---: | --- | --- | --- |",
    ]
    for item in audit.rubric:
        lines.append(
            f"| {item.item} | {item.max_score:.2f} | {item.score:.3f} | {item.status} | {escape_md(item.evidence)} | {escape_md('; '.join(item.flags))} |"
        )
    lines.extend(["", "## Observacoes", ""])
    for note in audit.notes:
        lines.append(f"- {note}")
    if audit.flags:
        lines.extend(["", "## Flags", ""])
        for flag in audit.flags:
            lines.append(f"- `{flag}`")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def escape_md(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ")


def write_group_feedback_html(audit: GroupAudit, path: Path) -> None:
    body = build_feedback_html_body(audit)
    path.write_text(f"<!DOCTYPE html><html><body>{body}</body></html>\n", encoding="utf-8")


def write_group_feedback_md(audit: GroupAudit, path: Path) -> None:
    positives = [item for item in audit.rubric if item.score > 0 and item.status not in {"pendente", "bloqueado"}]
    attention = [
        item
        for item in audit.rubric
        if item.flags or item.status in {"pendente", "parcial", "fragil", "nao_atendido", "bloqueado"}
    ]
    lines = [
        f"# Feedback Moodle - {audit.group}",
        "",
        f"**Nota sugerida:** {audit.local_score_rounded:.2f} / 5,00",
        "",
        f"**Status:** `{audit.final_status}`",
        "",
        "## Sintese",
        "",
        build_feedback_summary(audit),
        "",
        "## Evidencias Positivas",
        "",
    ]
    for item in positives:
        lines.append(f"- **{item.item}** ({item.score:.2f}/{item.max_score:.2f}): {item.evidence}")
    lines.extend(["", "## Pontos Que Justificam Desconto", ""])
    for item in attention:
        if item.score >= item.max_score and not item.flags:
            continue
        flag_text = f" Flags: {', '.join(item.flags)}." if item.flags else ""
        lines.append(f"- **{item.item}** ({item.score:.2f}/{item.max_score:.2f}): {item.evidence}.{flag_text}")
    lines.extend(["", "## Observacoes De Auditoria", ""])
    for note in audit.notes:
        lines.append(f"- {note}")
    if audit.flags:
        lines.extend(["", "## Flags", ""])
        for flag in audit.flags:
            lines.append(f"- `{flag}`")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_feedback_summary(audit: GroupAudit) -> str:
    if audit.group == "Grupo A":
        return (
            "O grupo apresentou estrutura completa de planilha com abas equivalentes as exigidas, calculos estatisticos "
            "majoritariamente verificaveis e boas interpretacoes univariadas. Os descontos principais decorrem de "
            "planejamento insuficiente no DreamShaper, link de CSV apontando para a propria planilha e necessidade de "
            "maior cautela na conclusao/limitacoes."
        )
    if audit.group == "Grupo B":
        return (
            "O grupo apresentou links acessiveis e estrutura de abas obrigatorias, com parte das frequencias e medidas "
            "quantitativas verificaveis. Os descontos principais decorrem de planejamento insuficiente no DreamShaper, "
            "inconsistencia entre o N=362 da base e o N=342 declarado na inspecao de qualidade, sintese/interpretações "
            "estatisticas insuficientes e uso incompleto de formulas estatisticas nativas."
        )
    return "Feedback gerado a partir da rubrica, auditorias estatisticas e auditorias interpretativas."


def build_feedback_html_body(audit: GroupAudit) -> str:
    strengths = []
    issues = []
    for item in audit.rubric:
        if item.score > 0 and item.status != "pendente":
            strengths.append(f"{item.item}: {item.evidence}")
        if item.status in {"pendente", "parcial", "fragil", "nao_atendido", "bloqueado"}:
            issues.append(f"{item.item}: {item.evidence}")
    flags = ", ".join(audit.flags) if audit.flags else "sem flags criticas"
    return (
        f"<h3>{html.escape(audit.group)} - T2 N724</h3>"
        "<p><strong>Status:</strong> avaliacao preliminar com revisao manual obrigatoria dos links e da planilha.</p>"
        f"<p><strong>Nota local preliminar:</strong> {audit.local_score_rounded:.2f} / 5,00. "
        "Esta nota nao deve ser lancada como final antes da verificacao da Google Planilha.</p>"
        "<p><strong>Evidencias positivas:</strong></p>"
        f"<ul>{''.join(f'<li>{html.escape(s)}</li>' for s in strengths[:5])}</ul>"
        "<p><strong>Pendencias principais:</strong></p>"
        f"<ul>{''.join(f'<li>{html.escape(i)}</li>' for i in issues[:8])}</ul>"
        f"<p><strong>Flags:</strong> {html.escape(flags)}</p>"
    )


def write_summary(audits: list[GroupAudit], path: Path) -> None:
    any_remote = any(audit.final_status == "preliminar_com_auditoria_remota" for audit in audits)
    intro = (
        "Esta saida inclui auditoria remota dos links acessiveis, mas continua preliminar porque itens qualitativos e graficos podem exigir revisao manual."
        if any_remote
        else "Esta saida e preliminar: os blocos de Google Planilha, CSV bruto e interpretacao dependem de verificacao dos links externos."
    )
    lines = [
        "# Resumo Da Correcao Local - T2 N724",
        "",
        intro,
        "",
        "| Grupo | Nota local arredondada | Status | Flags principais |",
        "| --- | ---: | --- | --- |",
    ]
    for audit in audits:
        lines.append(f"| {audit.group} | {audit.local_score_rounded:.2f} | {audit.final_status} | {escape_md('; '.join(audit.flags[:8]))} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def extract_dreamshaper_pdf(input_path: Path, out: Path, min_chars_page: int) -> int:
    out.mkdir(parents=True, exist_ok=True)
    pdfs = discover_pdfs(input_path)
    if not pdfs:
        print(f"Nenhum PDF encontrado em {input_path}")
        return 1

    reports: list[PdfExtractionReport] = []
    for pdf in pdfs:
        print(f"[correcao_t2_n724] extraindo {pdf}")
        report = extract_pdf(pdf, out, min_chars_page)
        reports.append(report)

    (out / "dreamshaper_pdf_diagnostico.json").write_text(
        json.dumps([report_to_json(r) for r in reports], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_index(out, reports)
    scanned = [r for r in reports if r.likely_scanned]
    if scanned:
        print("[correcao_t2_n724] atencao: ha PDF provavelmente escaneado; revisar OCR.")
    return 0


def discover_pdfs(input_path: Path) -> list[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".pdf":
        return [input_path]
    if input_path.is_dir():
        return sorted(input_path.glob("*.pdf"))
    return []


def extract_pdf(pdf: Path, out: Path, min_chars_page: int) -> PdfExtractionReport:
    safe_name = slugify(pdf.stem)
    md_path = out / f"{safe_name}.md"
    txt_path = out / f"{safe_name}.txt"

    md_blocks: list[str] = [f"# DreamShaper - {pdf.name}", "", f"Arquivo original: `{pdf}`", ""]
    txt_blocks: list[str] = []
    page_reports: list[PageExtraction] = []
    total_chars = 0
    total_words = 0
    total_tables = 0

    with pdfplumber.open(pdf) as document:
        for idx, page in enumerate(document.pages, start=1):
            text = (page.extract_text(x_tolerance=1, y_tolerance=3) or "").strip()
            tables = page.extract_tables() or []
            chars = len(text)
            words = len(re.findall(r"\S+", text))
            likely_image = chars < min_chars_page and not tables

            total_chars += chars
            total_words += words
            total_tables += len(tables)
            page_reports.append(
                PageExtraction(page=idx, chars=chars, words=words, tables=len(tables), likely_image=likely_image)
            )

            md_blocks.extend([f"## Pagina {idx}", ""])
            if text:
                md_blocks.extend([text, ""])
                txt_blocks.extend([f"--- Pagina {idx} ---", text, ""])
            else:
                md_blocks.extend(["> Sem texto extraido nesta pagina.", ""])
                txt_blocks.extend([f"--- Pagina {idx} ---", "Sem texto extraido nesta pagina.", ""])

            for table_idx, table in enumerate(tables, start=1):
                md_blocks.extend([f"### Tabela {table_idx}", ""])
                md_blocks.extend(table_to_markdown(table))
                md_blocks.append("")

    likely_scanned = bool(page_reports) and sum(p.likely_image for p in page_reports) / len(page_reports) >= 0.5
    if likely_scanned:
        md_blocks.extend(
            [
                "## Diagnostico",
                "",
                "Este PDF parece conter paginas em imagem ou com texto insuficiente. A correcao deve aplicar OCR ou revisao manual antes de usar o conteudo extraido.",
                "",
            ]
        )

    md_path.write_text("\n".join(md_blocks).strip() + "\n", encoding="utf-8")
    txt_path.write_text("\n".join(txt_blocks).strip() + "\n", encoding="utf-8")

    return PdfExtractionReport(
        file=str(pdf),
        pages=len(page_reports),
        chars=total_chars,
        words=total_words,
        tables=total_tables,
        likely_scanned=likely_scanned,
        output_md=str(md_path),
        output_txt=str(txt_path),
        page_reports=page_reports,
    )


def table_to_markdown(table: list[list[Any]]) -> list[str]:
    rows = [[clean_cell(cell) for cell in row] for row in table if row]
    if not rows:
        return []
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = normalized[0]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in normalized[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return lines


def clean_cell(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text.replace("|", "\\|")


def slugify(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    value = re.sub(r"_+", "_", value).strip("._-")
    return value or "dreamshaper"


def report_to_json(report: PdfExtractionReport) -> dict[str, Any]:
    data = asdict(report)
    data["page_reports"] = [asdict(page) for page in report.page_reports]
    return data


def write_index(out: Path, reports: list[PdfExtractionReport]) -> None:
    lines = ["# Diagnostico Da Extracao Dos PDFs Do DreamShaper", ""]
    for report in reports:
        lines.extend(
            [
                f"## {Path(report.file).name}",
                "",
                f"- Paginas: {report.pages}",
                f"- Caracteres extraidos: {report.chars}",
                f"- Palavras extraidas: {report.words}",
                f"- Tabelas detectadas: {report.tables}",
                f"- Provavel PDF escaneado: {'sim' if report.likely_scanned else 'nao'}",
                f"- Markdown: `{report.output_md}`",
                f"- Texto: `{report.output_txt}`",
                "",
            ]
        )
    (out / "dreamshaper_pdf_diagnostico.md").write_text("\n".join(lines), encoding="utf-8")
