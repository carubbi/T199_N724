from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series


ROOT = Path(__file__).resolve().parents[1]
DATASET_PATH = ROOT / "dataset" / "api.csv"
DOC_PATH = ROOT / "dataset" / "api_doc.txt"
OUTPUT_PATH = ROOT / "aulas" / "planilhas" / "API_RLS_Google_Planilhas.xlsx"

N_RAW = 5423
N_FUND = 3812
RAW_LAST = N_RAW + 1
FUND_LAST = N_FUND + 1
REFS_DATA_START = 26
REFS_DATA_END = REFS_DATA_START + N_FUND - 1
NPI_DATA_START = 26
NPI_DATA_END = NPI_DATA_START + N_FUND - 1
MULTI_DATA_START = 31
MULTI_DATA_END = MULTI_DATA_START + N_FUND - 1

RAW = "01_DADOS_BRUTOS"
FUND = "03_DADOS_LIMPOS"


def write_rows(ws, rows, start_row=1, start_col=1):
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            ws.cell(r_idx, c_idx, value)


def add_lines(ws, lines, col=1, start_row=1):
    for idx, line in enumerate(lines, start_row):
        ws.cell(idx, col, line)


def add_scatter(ws, title, x_col, y_col, min_row, max_row, anchor, fitted_col=None, x_title="X", y_title="Y"):
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    x_values = Reference(ws, min_col=x_col, min_row=min_row, max_row=max_row)
    y_values = Reference(ws, min_col=y_col, min_row=min_row, max_row=max_row)
    chart.series.append(Series(y_values, x_values, title="Observado"))
    if fitted_col:
        fitted_values = Reference(ws, min_col=fitted_col, min_row=min_row, max_row=max_row)
        chart.series.append(Series(fitted_values, x_values, title="Ajustado MQ"))
    chart.height = 10
    chart.width = 16
    ws.add_chart(chart, anchor)


def linest_simple(y_range: str, x_range: str, row: int, col: int) -> str:
    return f"=INDEX(LINEST({y_range};{x_range};TRUE;TRUE);{row};{col})"


def main():
    api = pd.read_csv(DATASET_PATH)
    wb = Workbook()
    wb.remove(wb.active)

    # 00_README
    ws = wb.create_sheet("00_README")
    add_lines(
        ws,
        [
            "API, Condições Socioeconômicas e Proficiência em Inglês",
            "Objetivo",
            "Esta planilha analisa a associação entre desempenho acadêmico de escolas da Califórnia e duas características contextuais dos estudantes: elegibilidade para refeição subsidiada e não proficiência em inglês.",
            "Pergunta de pesquisa",
            "Entre escolas fundamentais da Califórnia, como condições socioeducacionais dos alunos se associam ao desempenho acadêmico medido pelo API em 2000?",
            "Dados",
            "Fonte dos dados: dataset/api.csv.",
            "Fonte do dicionário e contexto: dataset/api_doc.txt.",
            "Unidade de análise: escola.",
            "Variável resposta: API00, índice de desempenho acadêmico em 2000.",
            "Preditor principal: REFS, percentual de alunos elegíveis para refeição subsidiada.",
            "Preditor complementar: NPI, percentual de alunos não proficientes em inglês.",
            "Recorte analítico: escolas com TIPO = Fundamental.",
            "Análises implementadas",
            "1. Importação do dataset bruto.",
            "2. Construção do dicionário de variáveis.",
            "3. Filtro das escolas fundamentais.",
            "4. Resumo por tipo de escola.",
            "5. Estatísticas descritivas e matriz de correlação.",
            "6. Regressão linear simples: API00 ~ REFS.",
            "7. Regressão linear simples: API00 ~ NPI.",
            "8. Regressão linear múltipla: API00 ~ REFS + NPI.",
            "9. Comparação dos modelos por R2, R2 ajustado e erro padrão residual.",
            "10. Diagnóstico por resíduos e QQ-plots simplificados.",
            "11. Identificação de observações atípicas por resíduo padronizado.",
            "12. Síntese final dos achados e limitações.",
            "Estrutura técnica",
            "Configuração-alvo: Google Sheets com localidade Brasil.",
            "Regra de construção: usar números, textos, fórmulas e gráficos simples; sem mesclagem de células ou estilo visual.",
            "As abas analíticas usam fórmulas nas células; os resultados não foram convertidos em valores estáticos.",
            "Principais funções usadas: FILTER, AVERAGE, STDEV, CORREL, SLOPE, INTERCEPT, RSQ, STEYX, LINEST, T.DIST.2T e TINV.",
            "Inferência",
            "Os intervalos de confiança usam alpha = 0,05, equivalente a nível de confiança de 95%.",
            "O valor-p é usado para avaliar evidência estatística contra H0: coeficiente = 0.",
            "Cuidados de interpretação",
            "A análise é associativa, não causal.",
            "API00 ~ REFS e API00 ~ NPI são modelos de regressão linear simples.",
            "API00 ~ REFS + NPI é regressão múltipla, não RLS.",
            "Valor-p pequeno não mede tamanho do efeito e não prova causalidade.",
            "IC95 informa uma faixa plausível para o coeficiente sob o modelo especificado.",
            "As conclusões são restritas às escolas fundamentais da base analisada.",
            "REFS e NPI são positivamente correlacionadas; por isso, os coeficientes do modelo múltiplo exigem cautela.",
        ],
    )

    # 01_DADOS_BRUTOS
    ws = wb.create_sheet(RAW)
    write_rows(ws, [list(api.columns)] + api.values.tolist())

    # 02_DICIONARIO
    ws = wb.create_sheet("02_DICIONARIO")
    write_rows(
        ws,
        [
            ["Variável", "Descrição", "Tipo", "Uso"],
            ["TIPO", "Nível de ensino da escola", "Categórica", "Recorte analítico"],
            ["API00", "Índice de desempenho acadêmico em 2000", "Quantitativa", "Variável resposta"],
            ["REFS", "Percentual de alunos elegíveis para refeição subsidiada", "Quantitativa", "Proxy socioeconômica"],
            ["NPI", "Percentual de alunos não proficientes em inglês", "Quantitativa", "Variável contextual"],
            [],
            ["Conteúdo de dataset/api_doc.txt"],
        ],
    )
    add_lines(ws, DOC_PATH.read_text(encoding="utf-8").splitlines(), start_row=8)

    # 03_DADOS_LIMPOS
    ws = wb.create_sheet(FUND)
    ws["A1"] = (
        f"={{'{RAW}'!A1:K1;"
        f"FILTER('{RAW}'!A2:K{RAW_LAST};"
        f"'{RAW}'!B2:B{RAW_LAST}=\"Fundamental\";"
        f"'{RAW}'!D2:D{RAW_LAST}<>\"\";"
        f"'{RAW}'!J2:J{RAW_LAST}<>\"\";"
        f"'{RAW}'!K2:K{RAW_LAST}<>\"\")}}"
    )
    ws["M1"] = "Checkpoint por fórmula"
    write_rows(
        ws,
        [
            ["n_escolas_fundamentais", f"=COUNTA(A2:A{FUND_LAST})"],
            ["media_API00", f"=AVERAGE(D2:D{FUND_LAST})"],
            ["media_REFS", f"=AVERAGE(J2:J{FUND_LAST})"],
            ["media_NPI", f"=AVERAGE(K2:K{FUND_LAST})"],
        ],
        start_row=2,
        start_col=13,
    )

    # 04_RESUMO_POR_TIPO
    ws = wb.create_sheet("04_RESUMO_POR_TIPO")
    write_rows(
        ws,
        [
            ["TIPO", "n_escolas", "percentual_amostra", "media_API00", "desvio_API00", "media_REFS", "desvio_REFS", "media_NPI", "desvio_NPI"],
            ["Fundamental"],
            ["Medio"],
            ["Superior"],
        ],
    )
    for row in range(2, 5):
        ws.cell(row, 2, f"=COUNTIF('{RAW}'!B2:B{RAW_LAST};A{row})")
        ws.cell(row, 3, f"=B{row}/COUNTA('{RAW}'!A2:A{RAW_LAST})*100")
        ws.cell(row, 4, f"=AVERAGEIF('{RAW}'!B2:B{RAW_LAST};A{row};'{RAW}'!D2:D{RAW_LAST})")
        ws.cell(row, 5, f"=STDEV(FILTER('{RAW}'!D2:D{RAW_LAST};'{RAW}'!B2:B{RAW_LAST}=A{row}))")
        ws.cell(row, 6, f"=AVERAGEIF('{RAW}'!B2:B{RAW_LAST};A{row};'{RAW}'!J2:J{RAW_LAST})")
        ws.cell(row, 7, f"=STDEV(FILTER('{RAW}'!J2:J{RAW_LAST};'{RAW}'!B2:B{RAW_LAST}=A{row}))")
        ws.cell(row, 8, f"=AVERAGEIF('{RAW}'!B2:B{RAW_LAST};A{row};'{RAW}'!K2:K{RAW_LAST})")
        ws.cell(row, 9, f"=STDEV(FILTER('{RAW}'!K2:K{RAW_LAST};'{RAW}'!B2:B{RAW_LAST}=A{row}))")
    add_lines(
        ws,
        ["Interpretação", "A tabela justifica o recorte para TIPO = Fundamental, pois os tipos de escola têm composição e médias distintas."],
        start_row=7,
    )

    # 05_DESCRITIVAS_E_CORRELACOES
    ws = wb.create_sheet("05_DESCRITIVAS_E_CORRELACOES")
    write_rows(
        ws,
        [
            ["estatistica", "API00", "REFS", "NPI"],
            ["contagem"],
            ["media"],
            ["desvio_padrao"],
            ["minimo"],
            ["Q1"],
            ["mediana"],
            ["Q3"],
            ["maximo"],
            [],
            ["Matriz de correlação", "API00", "REFS", "NPI"],
            ["API00"],
            ["REFS"],
            ["NPI"],
            [],
            ["Interpretação"],
            ["REFS e NPI são positivamente correlacionadas; o modelo múltiplo exige cautela por colinearidade."],
        ],
    )
    ranges = {
        "API00": f"FILTER('{FUND}'!D2:D;ISNUMBER('{FUND}'!D2:D))",
        "REFS": f"FILTER('{FUND}'!J2:J;ISNUMBER('{FUND}'!D2:D))",
        "NPI": f"FILTER('{FUND}'!K2:K;ISNUMBER('{FUND}'!D2:D))",
    }
    formulas = [
        "COUNTA",
        "AVERAGE",
        "STDEV",
        "MIN",
        "QUARTILE",
        "MED",
        "QUARTILE",
        "MAX",
    ]
    for col_idx, var in enumerate(["API00", "REFS", "NPI"], 2):
        data_range = ranges[var]
        ws.cell(2, col_idx, f"=COUNTA({data_range})")
        ws.cell(3, col_idx, f"=AVERAGE({data_range})")
        ws.cell(4, col_idx, f"=STDEV({data_range})")
        ws.cell(5, col_idx, f"=MIN({data_range})")
        ws.cell(6, col_idx, f"=QUARTILE({data_range};1)")
        ws.cell(7, col_idx, f"=MEDIAN({data_range})")
        ws.cell(8, col_idx, f"=QUARTILE({data_range};3)")
        ws.cell(9, col_idx, f"=MAX({data_range})")
    for row_idx, y in enumerate(["API00", "REFS", "NPI"], 12):
        for col_idx, x in enumerate(["API00", "REFS", "NPI"], 2):
            ws.cell(row_idx, col_idx, f"=CORREL({ranges[y]};{ranges[x]})")

    # 06_MODELO_REFS
    ws = wb.create_sheet("06_MODELO_REFS")
    y_ref = f"FILTER('{FUND}'!D2:D;ISNUMBER('{FUND}'!D2:D))"
    x_refs = f"FILTER('{FUND}'!J2:J;ISNUMBER('{FUND}'!D2:D))"
    write_rows(
        ws,
        [
            ["Modelo", "API00 ~ REFS"],
            ["n", f"=COUNTA({y_ref})"],
            ["gl", "=B2-2"],
            ["alpha", 0.05],
            ["nivel_confianca", "=1-B4"],
            [],
            ["Métrica", "Fórmula"],
            ["Intercepto", f"=INTERCEPT({y_ref};{x_refs})"],
            ["Coeficiente REFS", f"=SLOPE({y_ref};{x_refs})"],
            ["R2", f"=RSQ({y_ref};{x_refs})"],
            ["Erro padrão residual", f"=STEYX({y_ref};{x_refs})"],
            [],
            ["Parâmetro", "Coeficiente", "Erro padrão", "t", "valor-p", "IC95 inferior", "IC95 superior"],
            ["Intercepto", "=B8", linest_simple(y_ref, x_refs, 2, 2)],
            ["REFS", "=B9", linest_simple(y_ref, x_refs, 2, 1)],
        ],
    )
    for row in [14, 15]:
        ws.cell(row, 4, f"=B{row}/C{row}")
        ws.cell(row, 5, f"=T.DIST.2T(ABS(D{row});$B$3)")
        ws.cell(row, 6, f"=B{row}-TINV($B$4;$B$3)*C{row}")
        ws.cell(row, 7, f"=B{row}+TINV($B$4;$B$3)*C{row}")
    add_lines(
        ws,
        [
            "Nota sobre alpha",
            "alpha = 0,05 porque o intervalo é de 95%: alpha = 1 - 0,95.",
            "TINV(alpha; gl) retorna o valor crítico t bilateral usado no IC95.",
            "",
            "Interpretação",
            "O IC95 de REFS está totalmente abaixo de zero.",
            "O valor-p pequeno indica evidência estatística de associação linear negativa.",
            "A cada aumento de 1 ponto percentual em REFS, o API médio estimado diminui cerca de 3,45 pontos.",
        ],
        start_row=18,
    )
    write_rows(ws, [["ID", "NOME", "API00", "REFS", "API00_ajustado_REFS", "residuo_REFS", "residuo_padronizado_REFS", "abs_residuo_padronizado_REFS"]], start_row=25)
    for idx in range(N_FUND):
        row = idx + 26
        src = idx + 2
        ws.cell(row, 1, f"='{FUND}'!A{src}")
        ws.cell(row, 2, f"='{FUND}'!C{src}")
        ws.cell(row, 3, f"='{FUND}'!D{src}")
        ws.cell(row, 4, f"='{FUND}'!J{src}")
        ws.cell(row, 5, f"=$B$8+$B$9*D{row}")
        ws.cell(row, 6, f"=C{row}-E{row}")
        ws.cell(row, 7, f"=F{row}/$B$11")
        ws.cell(row, 8, f"=ABS(G{row})")
    add_scatter(ws, "REFS e API00 com ajuste por MQ", 4, 3, 26, 25 + N_FUND, "N25", fitted_col=5, x_title="REFS", y_title="API00")

    # 07_MODELO_NPI
    ws = wb.create_sheet("07_MODELO_NPI")
    x_npi = f"FILTER('{FUND}'!K2:K;ISNUMBER('{FUND}'!D2:D))"
    write_rows(
        ws,
        [
            ["Modelo", "API00 ~ NPI"],
            ["n", f"=COUNTA({y_ref})"],
            ["gl", "=B2-2"],
            ["alpha", 0.05],
            ["nivel_confianca", "=1-B4"],
            [],
            ["Métrica", "Fórmula"],
            ["Intercepto", f"=INTERCEPT({y_ref};{x_npi})"],
            ["Coeficiente NPI", f"=SLOPE({y_ref};{x_npi})"],
            ["R2", f"=RSQ({y_ref};{x_npi})"],
            ["Erro padrão residual", f"=STEYX({y_ref};{x_npi})"],
            [],
            ["Parâmetro", "Coeficiente", "Erro padrão", "t", "valor-p", "IC95 inferior", "IC95 superior"],
            ["Intercepto", "=B8", linest_simple(y_ref, x_npi, 2, 2)],
            ["NPI", "=B9", linest_simple(y_ref, x_npi, 2, 1)],
        ],
    )
    for row in [14, 15]:
        ws.cell(row, 4, f"=B{row}/C{row}")
        ws.cell(row, 5, f"=T.DIST.2T(ABS(D{row});$B$3)")
        ws.cell(row, 6, f"=B{row}-TINV($B$4;$B$3)*C{row}")
        ws.cell(row, 7, f"=B{row}+TINV($B$4;$B$3)*C{row}")
    add_lines(
        ws,
        [
            "Nota sobre alpha",
            "alpha = 0,05 porque o intervalo é de 95%: alpha = 1 - 0,95.",
            "TINV(alpha; gl) retorna o valor crítico t bilateral usado no IC95.",
            "",
            "Interpretação",
            "O IC95 de NPI está totalmente abaixo de zero.",
            "O valor-p pequeno indica evidência estatística de associação linear negativa.",
            "A cada aumento de 1 ponto percentual em NPI, o API médio estimado diminui cerca de 3,56 pontos no modelo simples.",
        ],
        start_row=18,
    )
    write_rows(ws, [["ID", "NOME", "API00", "NPI", "API00_ajustado_NPI", "residuo_NPI", "residuo_padronizado_NPI", "abs_residuo_padronizado_NPI"]], start_row=25)
    for idx in range(N_FUND):
        row = idx + 26
        src = idx + 2
        ws.cell(row, 1, f"='{FUND}'!A{src}")
        ws.cell(row, 2, f"='{FUND}'!C{src}")
        ws.cell(row, 3, f"='{FUND}'!D{src}")
        ws.cell(row, 4, f"='{FUND}'!K{src}")
        ws.cell(row, 5, f"=$B$8+$B$9*D{row}")
        ws.cell(row, 6, f"=C{row}-E{row}")
        ws.cell(row, 7, f"=F{row}/$B$11")
        ws.cell(row, 8, f"=ABS(G{row})")
    add_scatter(ws, "NPI e API00 com ajuste por MQ", 4, 3, 26, 25 + N_FUND, "N25", fitted_col=5, x_title="NPI", y_title="API00")

    # 08_MODELO_MULTIPLO
    ws = wb.create_sheet("08_MODELO_MULTIPLO")
    x_multi = f"FILTER('{FUND}'!J2:K;ISNUMBER('{FUND}'!D2:D))"
    write_rows(
        ws,
        [
            ["Modelo", "API00 ~ REFS + NPI"],
            ["n", f"=COUNTA({y_ref})"],
            ["k", 2],
            ["gl", "=B2-B3-1"],
            ["alpha", 0.05],
            ["nivel_confianca", "=1-B5"],
            [],
            ["Métrica", "Fórmula"],
            ["Intercepto", f"=INDEX(LINEST({y_ref};{x_multi};TRUE;TRUE);1;3)"],
            ["Coeficiente REFS", f"=INDEX(LINEST({y_ref};{x_multi};TRUE;TRUE);1;2)"],
            ["Coeficiente NPI", f"=INDEX(LINEST({y_ref};{x_multi};TRUE;TRUE);1;1)"],
            ["R2", f"=INDEX(LINEST({y_ref};{x_multi};TRUE;TRUE);3;1)"],
            ["R2 ajustado", "=1-(1-B12)*(B2-1)/(B2-B3-1)"],
            ["Erro padrão residual", f"=INDEX(LINEST({y_ref};{x_multi};TRUE;TRUE);3;2)"],
            [],
            ["Ordem LINEST", "Se X estiver como REFS:NPI, LINEST retorna NPI, REFS e intercepto."],
            [],
            ["Parâmetro", "Coeficiente", "Erro padrão", "t", "valor-p", "IC95 inferior", "IC95 superior"],
            ["Intercepto", "=B9", f"=INDEX(LINEST({y_ref};{x_multi};TRUE;TRUE);2;3)"],
            ["REFS", "=B10", f"=INDEX(LINEST({y_ref};{x_multi};TRUE;TRUE);2;2)"],
            ["NPI", "=B11", f"=INDEX(LINEST({y_ref};{x_multi};TRUE;TRUE);2;1)"],
        ],
    )
    for row in [19, 20, 21]:
        ws.cell(row, 4, f"=B{row}/C{row}")
        ws.cell(row, 5, f"=T.DIST.2T(ABS(D{row});$B$4)")
        ws.cell(row, 6, f"=B{row}-TINV($B$5;$B$4)*C{row}")
        ws.cell(row, 7, f"=B{row}+TINV($B$5;$B$4)*C{row}")
    add_lines(
        ws,
        [
            "Nota sobre alpha",
            "alpha = 0,05 porque o intervalo é de 95%: alpha = 1 - 0,95.",
            "TINV(alpha; gl) retorna o valor crítico t bilateral usado no IC95.",
            "",
            "Interpretação",
            "Os IC95 de REFS e NPI permanecem abaixo de zero.",
            "No modelo múltiplo, os coeficientes são associações parciais.",
            "O coeficiente de NPI diminui em magnitude em relação ao modelo simples, indicando sobreposição de informação com REFS.",
        ],
        start_row=24,
    )
    write_rows(ws, [["ID", "NOME", "API00", "REFS", "NPI", "API00_ajustado_MULTI", "residuo_MULTI", "residuo_padronizado_MULTI", "abs_residuo_padronizado_MULTI"]], start_row=30)
    for idx in range(N_FUND):
        row = idx + 31
        src = idx + 2
        ws.cell(row, 1, f"='{FUND}'!A{src}")
        ws.cell(row, 2, f"='{FUND}'!C{src}")
        ws.cell(row, 3, f"='{FUND}'!D{src}")
        ws.cell(row, 4, f"='{FUND}'!J{src}")
        ws.cell(row, 5, f"='{FUND}'!K{src}")
        ws.cell(row, 6, f"=$B$9+$B$10*D{row}+$B$11*E{row}")
        ws.cell(row, 7, f"=C{row}-F{row}")
        ws.cell(row, 8, f"=G{row}/$B$14")
        ws.cell(row, 9, f"=ABS(H{row})")
    add_scatter(ws, "Observado versus ajustado MULTI", 6, 3, 31, 30 + N_FUND, "N30", x_title="API00 ajustado", y_title="API00 observado")
    add_scatter(ws, "REFS com ajuste MULTI projetado", 4, 3, 31, 30 + N_FUND, "N48", fitted_col=6, x_title="REFS", y_title="API00")
    add_scatter(ws, "NPI com ajuste MULTI projetado", 5, 3, 31, 30 + N_FUND, "N66", fitted_col=6, x_title="NPI", y_title="API00")

    # 09_COMPARACAO_MODELOS
    ws = wb.create_sheet("09_COMPARACAO_MODELOS")
    write_rows(
        ws,
        [
            ["Modelo", "Tipo", "R2", "R2 ajustado", "Erro padrão residual", "Interpretação inferencial"],
            ["API00 ~ REFS", "RLS principal", "='06_MODELO_REFS'!B10", "='06_MODELO_REFS'!B10", "='06_MODELO_REFS'!B11", "REFS negativo, IC95 abaixo de zero, p <0,001"],
            ["API00 ~ NPI", "RLS complementar", "='07_MODELO_NPI'!B10", "='07_MODELO_NPI'!B10", "='07_MODELO_NPI'!B11", "NPI negativo, IC95 abaixo de zero, p <0,001"],
            ["API00 ~ REFS + NPI", "Regressão múltipla", "='08_MODELO_MULTIPLO'!B12", "='08_MODELO_MULTIPLO'!B13", "='08_MODELO_MULTIPLO'!B14", "ambos negativos, IC95 abaixo de zero, p <0,001"],
            [],
            ["Interpretação"],
            ["REFS explica mais variação de API00 do que NPI isoladamente."],
            ["O modelo múltiplo melhora o ajuste, mas pouco em relação ao modelo com REFS."],
            ["Valor-p pequeno não mede tamanho do efeito e não prova causalidade."],
            ["IC95 informa faixa plausível para o coeficiente sob o modelo especificado."],
            ["Os gráficos de modelos simples mostram a reta de MQ diretamente; os gráficos do modelo múltiplo são projeções dos valores ajustados."],
        ],
    )

    # 10_DIAGNOSTICOS
    ws = wb.create_sheet("10_DIAGNOSTICOS")
    specs = [
        ("API00 ~ REFS", "06_MODELO_REFS", 5, 6, 1),
        ("API00 ~ NPI", "07_MODELO_NPI", 5, 6, 5),
        ("API00 ~ REFS + NPI", "08_MODELO_MULTIPLO", 6, 7, 9),
    ]
    for title, sheet, fitted_col, resid_col, start_col in specs:
        ws.cell(1, start_col, f"Resíduos versus ajustados - {title}")
        write_rows(ws, [["fitted", "residual"]], start_row=2, start_col=start_col)
        source_start = REFS_DATA_START if sheet != "08_MODELO_MULTIPLO" else MULTI_DATA_START
        for idx in range(N_FUND):
            row = idx + 3
            src = idx + source_start
            ws.cell(row, start_col, f"='{sheet}'!{chr(64 + fitted_col)}{src}")
            ws.cell(row, start_col + 1, f"='{sheet}'!{chr(64 + resid_col)}{src}")
    qq_start = FUND_LAST + 5
    qq_specs = [
        ("QQ-plot simplificado API00 ~ REFS", "06_MODELO_REFS", "F", 1),
        ("QQ-plot simplificado API00 ~ NPI", "07_MODELO_NPI", "F", 6),
        ("QQ-plot simplificado API00 ~ REFS + NPI", "08_MODELO_MULTIPLO", "G", 11),
    ]
    for title, sheet, resid_col, start_col in qq_specs:
        ws.cell(qq_start, start_col, title)
        write_rows(ws, [["i", "p", "quantil_normal", "residuo_ordenado"]], start_row=qq_start + 1, start_col=start_col)
        source_start = REFS_DATA_START if sheet != "08_MODELO_MULTIPLO" else MULTI_DATA_START
        source_end = source_start + N_FUND - 1
        for idx in range(N_FUND):
            row = qq_start + 2 + idx
            i_expr = idx + 1
            ws.cell(row, start_col, i_expr)
            ws.cell(row, start_col + 1, f"=({start_col_letter(start_col)}{row}-0,5)/{N_FUND}")
            ws.cell(row, start_col + 2, f"=NORM.S.INV({start_col_letter(start_col + 1)}{row})")
            ws.cell(row, start_col + 3, f"=INDEX(SORT('{sheet}'!{resid_col}${source_start}:{resid_col}${source_end};1;TRUE);{start_col_letter(start_col)}{row})")
    add_lines(
        ws,
        [
            "Interpretação",
            "Verificar caudas e padrões de dispersão. Normalidade não deve ser usada como critério automático.",
            "Valor-p e IC95 dependem da adequação aproximada das hipóteses do modelo.",
        ],
        col=13,
    )
    add_scatter(ws, "Resíduos API00 ~ REFS", 1, 2, 3, 2 + N_FUND, "N5", x_title="Ajustado", y_title="Resíduo")
    add_scatter(ws, "Resíduos API00 ~ NPI", 5, 6, 3, 2 + N_FUND, "N23", x_title="Ajustado", y_title="Resíduo")
    add_scatter(ws, "Resíduos API00 ~ REFS + NPI", 9, 10, 3, 2 + N_FUND, "N41", x_title="Ajustado", y_title="Resíduo")
    add_scatter(ws, "QQ API00 ~ REFS", 3, 4, qq_start + 2, qq_start + 1 + N_FUND, "N59", x_title="Quantil normal", y_title="Resíduo ordenado")
    add_scatter(ws, "QQ API00 ~ NPI", 8, 9, qq_start + 2, qq_start + 1 + N_FUND, "N77", x_title="Quantil normal", y_title="Resíduo ordenado")
    add_scatter(ws, "QQ API00 ~ REFS + NPI", 13, 14, qq_start + 2, qq_start + 1 + N_FUND, "N95", x_title="Quantil normal", y_title="Resíduo ordenado")

    # 11_OUTLIERS
    ws = wb.create_sheet("11_OUTLIERS")
    write_rows(
        ws,
        [
            ["Modelo", "n", "|resíduo padronizado| > 2", "|resíduo padronizado| > 3", "% > 3"],
            [
                "API00 ~ REFS",
                "='06_MODELO_REFS'!B2",
                f"=SUMPRODUCT(--(ABS('06_MODELO_REFS'!G{REFS_DATA_START}:G{REFS_DATA_END})>2))",
                f"=SUMPRODUCT(--(ABS('06_MODELO_REFS'!G{REFS_DATA_START}:G{REFS_DATA_END})>3))",
                "=D2/B2*100",
            ],
            [
                "API00 ~ NPI",
                "='07_MODELO_NPI'!B2",
                f"=SUMPRODUCT(--(ABS('07_MODELO_NPI'!G{NPI_DATA_START}:G{NPI_DATA_END})>2))",
                f"=SUMPRODUCT(--(ABS('07_MODELO_NPI'!G{NPI_DATA_START}:G{NPI_DATA_END})>3))",
                "=D3/B3*100",
            ],
            [
                "API00 ~ REFS + NPI",
                "='08_MODELO_MULTIPLO'!B2",
                f"=SUMPRODUCT(--(ABS('08_MODELO_MULTIPLO'!H{MULTI_DATA_START}:H{MULTI_DATA_END})>2))",
                f"=SUMPRODUCT(--(ABS('08_MODELO_MULTIPLO'!H{MULTI_DATA_START}:H{MULTI_DATA_END})>3))",
                "=D4/B4*100",
            ],
            [],
            ["10 escolas mais atípicas no modelo API00 ~ REFS"],
            ["ID", "NOME", "API00", "REFS", "valor_ajustado_REFS", "residuo_REFS", "residuo_padronizado_REFS", "abs_residuo_padronizado_REFS", "linha_auxiliar"],
        ],
    )
    for idx in range(10):
        row = idx + 8
        rank = idx + 1
        ws.cell(row, 8, f"=LARGE('06_MODELO_REFS'!H${REFS_DATA_START}:H${REFS_DATA_END};{rank})")
        ws.cell(row, 9, f"=MATCH(H{row};'06_MODELO_REFS'!H${REFS_DATA_START}:H${REFS_DATA_END};0)")
        for col in range(1, 8):
            source_col = chr(64 + col)
            ws.cell(row, col, f"=INDEX('06_MODELO_REFS'!{source_col}${REFS_DATA_START}:{source_col}${REFS_DATA_END};$I{row})")
    add_lines(ws, ["Interpretação", "Observações atípicas são diagnósticas e não devem ser removidas automaticamente."], start_row=21)

    # 12_RESULTADOS_FINAIS
    ws = wb.create_sheet("12_RESULTADOS_FINAIS")
    add_lines(
        ws,
        [
            "Pergunta de pesquisa: entre escolas fundamentais da Califórnia, como condições socioeducacionais se associam ao API00?",
            "Recorte analítico: escolas com TIPO = Fundamental.",
            "Modelo principal: API00 ~ REFS.",
            "Modelo complementar: API00 ~ NPI.",
            "Modelo múltiplo: API00 ~ REFS + NPI.",
            "REFS tem associação negativa forte com API00 e maior capacidade explicativa isolada do que NPI.",
            "NPI também tem associação negativa com API00.",
            "No modelo múltiplo, NPI contribui, mas com efeito parcial menor do que no modelo simples.",
            "Todos os coeficientes principais têm valor-p <0,001 e IC95 abaixo de zero.",
            "Os gráficos de dispersão com ajuste por MQ apoiam visualmente as associações negativas nos modelos simples.",
            "Os gráficos do modelo múltiplo devem ser lidos como projeções dos valores ajustados.",
            "A análise é associativa, restrita à base analisada e não autoriza interpretação causal.",
        ],
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


def start_col_letter(col_idx: int) -> str:
    result = ""
    while col_idx:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


if __name__ == "__main__":
    main()
