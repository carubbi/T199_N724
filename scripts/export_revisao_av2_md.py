from __future__ import annotations

import math
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "aulas" / "planilhas" / "Revisao_AV2.xlsx"
EXPORT_DIR = ROOT / "aulas" / "planilhas" / "exports"


PT_BR_REPLACEMENTS = [
    ("Observacao", "Observação"),
    ("observacao", "observação"),
    ("variavel", "variável"),
    ("Variaveis", "Variáveis"),
    ("variaveis", "variáveis"),
    ("explicativa", "explicativa"),
    ("resposta", "resposta"),
    ("calculos", "cálculos"),
    ("Calculo", "Cálculo"),
    ("calculo", "cálculo"),
    ("media", "média"),
    ("desvio-padrao", "desvio-padrão"),
    ("determinacao", "determinação"),
    ("decomposicao", "decomposição"),
    ("correlacao", "correlação"),
    ("associacao", "associação"),
    ("relacao", "relação"),
    ("regressao", "regressão"),
    ("Regressao", "Regressão"),
    ("inclinacao", "inclinação"),
    ("interpretacao", "interpretação"),
    ("Interpretacao", "Interpretação"),
    ("previsao", "previsão"),
    ("exclusao", "exclusão"),
    ("decisao", "decisão"),
    ("Conferencia", "Conferência"),
    ("diferenca", "diferença"),
    ("aparencia", "aparência"),
    ("atipico", "atípico"),
    ("Atipico", "Atípico"),
    ("atipica", "atípica"),
    ("Atipica", "Atípica"),
    ("estatistica", "estatística"),
    ("automatica", "automática"),
    ("tambem", "também"),
    ("existencia", "existência"),
    ("critica", "crítica"),
    ("funcao", "função"),
    ("Funcao", "Função"),
    ("funcoes", "funções"),
    ("funcoes prontas", "funções prontas"),
    ("formula", "fórmula"),
    ("Formula", "Fórmula"),
    ("Substituicao", "Substituição"),
    ("exercicio", "exercício"),
    ("didatico", "didático"),
    ("numero", "número"),
    ("instituicao", "instituição"),
    ("instituicoes", "instituições"),
    ("residuos", "resíduos"),
    ("Residuos", "Resíduos"),
    ("proporcao", "proporção"),
    ("variacao", "variação"),
    ("contratacao", "contratação"),
    ("aceitar", "aceitar"),
    ("causalidade", "causalidade"),
    ("nao", "não"),
    ("Ha ", "Há "),
    ("Ha?", "Há?"),
    ("Grafico", "Gráfico"),
    ("grafico", "gráfico"),
    ("direcao", "direção"),
    ("cenario", "cenário"),
    ("apos", "após"),
    ("esta associado", "está associado"),
    ("estao associados", "estão associados"),
    ("associados", "associados"),
    ("tecnico", "técnico"),
    ("parametro", "parâmetro"),
    ("alem", "além"),
    ("Qual e ", "Qual é "),
    ("objetivo e ", "objetivo é "),
    ("regressão e pequena", "regressão é pequena"),
    ("regressão e grande", "regressão é grande"),
    ("reta e grande", "reta é grande"),
    ("relação a variação", "relação à variação"),
    ("comparada a variação", "comparada à variação"),
    ("intercepto e parâmetro", "intercepto é parâmetro"),
    ("relação a ponto", "relação a ponto"),
    ("PUC de Sao Paulo", "PUC de São Paulo"),
    ("Estadual de Sao Paulo", "Estadual de São Paulo"),
    ("Federal de Sao Carlos", "Federal de São Carlos"),
    ("Vicosa", "Viçosa"),
    ("Uberlandia", "Uberlândia"),
    ("Catolica", "Católica"),
    ("Petropolis", "Petrópolis"),
    ("Julio", "Júlio"),
    ("Andre", "André"),
]


def pt_br_text(value: object) -> str:
    text = "" if value is None else str(value)
    for old, new in PT_BR_REPLACEMENTS:
        text = text.replace(old, new)
    return text


def md_escape(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("|", "\\|")
    return text.replace("\n", "<br>")


def fmt(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if abs(value) >= 100:
            return f"{value:.3f}".rstrip("0").rstrip(".")
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value)


def fmt_br(value: object) -> str:
    return fmt(value).replace(".", ",")


def slug(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def normalize_md_table_line(line: str) -> str:
    cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
    return "| " + " | ".join(cells) + " |"


def alpha_item(line: str) -> str | None:
    match = re.match(r"^\s*([a-z])\)\s+(.+)$", line)
    if not match:
        return None
    return match.group(2).strip()


def format_enunciado(text: object) -> str:
    if text is None:
        return ""
    lines = pt_br_text(text).replace("\r\n", "\n").replace("\r", "\n").split("\n")
    out: list[str] = []
    in_table = False
    in_alpha_list = False
    for line in lines:
        is_table_line = "|" in line and line.strip()
        item = alpha_item(line)
        if is_table_line:
            if in_alpha_list:
                out.append("</ol>")
                in_alpha_list = False
            if not in_table:
                in_table = True
                out.append(normalize_md_table_line(line))
                columns = len(line.strip().strip("|").split("|"))
                out.append("| " + " | ".join(["---"] * columns) + " |")
            else:
                out.append(normalize_md_table_line(line))
        elif item is not None:
            in_table = False
            if not in_alpha_list:
                out.append('<ol type="a">')
                in_alpha_list = True
            out.append(f"<li>{item}</li>")
        else:
            if in_alpha_list:
                out.append("</ol>")
                in_alpha_list = False
            in_table = False
            stripped = line.strip()
            if stripped.startswith("Calcule"):
                out.append(f"**{stripped}**")
            else:
                out.append(line)
    if in_alpha_list:
        out.append("</ol>")
    return "\n".join(out).strip()


def find_row(ws, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == label:
            return row
    raise ValueError(f"Label not found in {ws.title}: {label}")


def get_data_rows(ws) -> list[tuple[str, float, float]]:
    start = 10
    end = find_row(ws, "Resumo do calculo manual") - 2
    rows = []
    for row in range(start, end + 1):
        name = ws.cell(row, 1).value
        x = ws.cell(row, 2).value
        y = ws.cell(row, 3).value
        if name is not None and x is not None and y is not None:
            rows.append((str(name), float(x), float(y)))
    return rows


def regression(rows: list[tuple[str, float, float]]) -> dict[str, float]:
    xs = [row[1] for row in rows]
    ys = [row[2] for row in rows]
    n = len(rows)
    mean_x = sum(xs) / n
    mean_y = sum(ys) / n
    dx = [x - mean_x for x in xs]
    dy = [y - mean_y for y in ys]
    sxy = sum(a * b for a, b in zip(dx, dy))
    sxx = sum(a * a for a in dx)
    syy = sum(b * b for b in dy)
    r = sxy / math.sqrt(sxx * syy)
    r2 = r * r
    b = sxy / sxx
    a = mean_y - b * mean_x
    yhat = [a + b * x for x in xs]
    residuals = [y - yh for y, yh in zip(ys, yhat)]
    sqres = sum(e * e for e in residuals)
    sqreg = sum((yh - mean_y) ** 2 for yh in yhat)
    sqtot = syy
    gl_reg = 1
    gl_res = n - 2
    gl_tot = n - 1
    qm_reg = sqreg / gl_reg
    qm_res = sqres / gl_res
    f_value = qm_reg / qm_res
    se_y = math.sqrt(qm_res)
    se_b = math.sqrt(qm_res / sxx)
    se_a = math.sqrt(qm_res * (1 / n + mean_x * mean_x / sxx))
    return {
        "n": n,
        "media x": mean_x,
        "media y": mean_y,
        "SOMA dx*dy": sxy,
        "SOMA dx^2": sxx,
        "SQTot = SOMA dy^2": sqtot,
        "r manual": r,
        "R2 manual = r^2": r2,
        "b manual": b,
        "a manual": a,
        "erro-padrao b": se_b,
        "erro-padrao a": se_a,
        "erro-padrao y": se_y,
        "Regressao SQ": sqreg,
        "Regressao gl": gl_reg,
        "Regressao QM": qm_reg,
        "Regressao F": f_value,
        "Residuos SQ": sqres,
        "Residuos gl": gl_res,
        "Residuos QM": qm_res,
        "Total SQ": sqtot,
        "Total gl": gl_tot,
        "Conferencia SQTot - (SQReg + SQRes)": sqtot - (sqreg + sqres),
        "R2 pela ANOVA": sqreg / sqtot,
    }


def lines_until_blank(ws, start_row: int) -> list[str]:
    lines = []
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if value is None:
            break
        lines.append(str(value))
    return lines


def summary_table(ws, stats: dict[str, float]) -> list[str]:
    start = find_row(ws, "Resumo do calculo manual") + 1
    end = find_row(ws, "ANOVA da regressao") - 2
    out = [
        "### Cálculo na planilha por somatórios",
        "",
        "| Medida | Fórmula na planilha | Resultado |",
        "| --- | --- | ---: |",
    ]
    for row in range(start, end + 1):
        label = ws.cell(row, 1).value
        formula = ws.cell(row, 2).value
        if label:
            out.append(f"| {md_escape(pt_br_text(label))} | `{md_escape(formula)}` | {fmt(stats.get(str(label)))} |")
    return out


def anova_table(ws, stats: dict[str, float]) -> list[str]:
    start = find_row(ws, "ANOVA da regressao")
    out = [
        "### ANOVA da regressão na planilha",
        "",
        "| Fonte | SQ | gl | QM | F |",
        "| --- | ---: | ---: | ---: | ---: |",
        (
            f"| Regressão | {fmt(stats['Regressao SQ'])} | {fmt(stats['Regressao gl'])} | "
            f"{fmt(stats['Regressao QM'])} | {fmt(stats['Regressao F'])} |"
        ),
        (
            f"| Resíduos | {fmt(stats['Residuos SQ'])} | {fmt(stats['Residuos gl'])} | "
            f"{fmt(stats['Residuos QM'])} |  |"
        ),
        f"| Total | {fmt(stats['Total SQ'])} | {fmt(stats['Total gl'])} |  |  |",
        "",
    ]
    conf_label = ws.cell(start + 6, 1).value
    r2_label = ws.cell(start + 7, 1).value
    out.extend(
        [
            f"- {pt_br_text(conf_label)}: {fmt(stats['Conferencia SQTot - (SQReg + SQRes)'])}",
            f"- {pt_br_text(r2_label)}: {fmt(stats['R2 pela ANOVA'])}",
        ]
    )
    return out


def validation_table(ws, stats: dict[str, float]) -> list[str]:
    end_row = find_row(ws, "Resumo do calculo manual") - 2
    y_range = f"C10:C{end_row}"
    x_range = f"B10:B{end_row}"
    out = [
        "### Validação com LINEST completo da planilha",
        "",
        "A função `LINEST` ajusta a regressão linear aos intervalos informados. O primeiro `TRUE` estima o intercepto da reta e o segundo `TRUE` retorna a matriz completa de estatísticas.",
        "",
        "| Modelo da fórmula | Substituição no exercício |",
        "| --- | --- |",
        f"| `=LINEST(intervalo_y; intervalo_x; TRUE; TRUE)` | `=LINEST({y_range}; {x_range}; TRUE; TRUE)` |",
        "",
        "| Linha do retorno | Coluna 1 | Resultado | Coluna 2 | Resultado |",
        "| --- | --- | ---: | --- | ---: |",
        f"| 1 | inclinação `b` | {fmt(stats['b manual'])} | intercepto `a` | {fmt(stats['a manual'])} |",
        f"| 2 | erro-padrão de `b` | {fmt(stats['erro-padrao b'])} | erro-padrão de `a` | {fmt(stats['erro-padrao a'])} |",
        f"| 3 | `R2` | {fmt(stats['R2 manual = r^2'])} | erro-padrão da estimativa de `y` | {fmt(stats['erro-padrao y'])} |",
        f"| 4 | `F` | {fmt(stats['Regressao F'])} | graus de liberdade | {fmt(stats['Residuos gl'])} |",
        f"| 5 | `SQReg` | {fmt(stats['Regressao SQ'])} | `SQRes` | {fmt(stats['Residuos SQ'])} |",
    ]
    return out


def data_table(ws, rows: list[tuple[str, float, float]]) -> list[str]:
    x_label = ws.cell(9, 2).value
    y_label = ws.cell(9, 3).value
    out = [
        "## Dados",
        "",
        f"| Caso | {md_escape(x_label)} | {md_escape(y_label)} |",
        "| --- | ---: | ---: |",
    ]
    for name, x, y in rows:
        out.append(f"| {md_escape(pt_br_text(name))} | {fmt(x)} | {fmt(y)} |")
    return out


def paper_calculation_section(rows: list[tuple[str, float, float]], stats: dict[str, float]) -> list[str]:
    sum_x = sum(row[1] for row in rows)
    sum_y = sum(row[2] for row in rows)
    n = int(stats["n"])
    gl_reg = int(stats["Regressao gl"])
    gl_res = int(stats["Residuos gl"])
    gl_tot = int(stats["Total gl"])

    aux_rows = [
        "| Caso | $x_i$ | $y_i$ | $d_{x_i}$ | $d_{y_i}$ | $d_{x_i}d_{y_i}$ | $d_{x_i}^2$ | $d_{y_i}^2$ |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for name, x, y in rows:
        dx = x - stats["media x"]
        dy = y - stats["media y"]
        aux_rows.append(
            "| "
            + " | ".join(
                [
                    md_escape(pt_br_text(name)),
                    fmt_br(x),
                    fmt_br(y),
                    fmt_br(dx),
                    fmt_br(dy),
                    fmt_br(dx * dy),
                    fmt_br(dx * dx),
                    fmt_br(dy * dy),
                ]
            )
            + " |"
        )
    aux_rows.extend(
        [
            "| **Somatório** | "
            + f"$\\sum x_i = {fmt_br(sum_x)}$ | "
            + f"$\\sum y_i = {fmt_br(sum_y)}$ |  |  | "
            + f"$\\sum d_{{x_i}}d_{{y_i}} = {fmt_br(stats['SOMA dx*dy'])}$ | "
            + f"$\\sum d_{{x_i}}^2 = {fmt_br(stats['SOMA dx^2'])}$ | "
            + f"$\\sum d_{{y_i}}^2 = {fmt_br(stats['SQTot = SOMA dy^2'])}$ |",
        ]
    )

    return [
        "## Desenvolvimento da solução sem o uso de planilha",
        "",
        "Esta seção mostra a resolução pelas fórmulas estatísticas, usando somatórios e substituição numérica. A planilha serve apenas para conferir os resultados.",
        "",
        "### 1. Médias",
        "",
        "$$",
        f"\\bar{{x}} = \\frac{{\\sum x_i}}{{n}} = \\frac{{{fmt_br(sum_x)}}}{{{n}}} = {fmt_br(stats['media x'])}",
        "$$",
        "",
        "$$",
        f"\\bar{{y}} = \\frac{{\\sum y_i}}{{n}} = \\frac{{{fmt_br(sum_y)}}}{{{n}}} = {fmt_br(stats['media y'])}",
        "$$",
        "",
        "### 2. Somatórios auxiliares",
        "",
        "Depois das médias, calcula-se para cada observação:",
        "",
        "$$",
        "d_{x_i} = x_i - \\bar{x}",
        "$$",
        "",
        "$$",
        "d_{y_i} = y_i - \\bar{y}",
        "$$",
        "",
        "A tabela a seguir explicita os produtos termo a termo e os somatórios usados nos cálculos.",
        "",
        *aux_rows,
        "",
        "Os somatórios principais são:",
        "",
        "| Somatório | Valor |",
        "| --- | ---: |",
        f"| $\\sum d_{{x_i}}d_{{y_i}}$ | {fmt_br(stats['SOMA dx*dy'])} |",
        f"| $\\sum d_{{x_i}}^2$ | {fmt_br(stats['SOMA dx^2'])} |",
        f"| $\\sum d_{{y_i}}^2$ | {fmt_br(stats['SQTot = SOMA dy^2'])} |",
        "",
        "### 3. Correlação de Pearson",
        "",
        "$$",
        "r = \\frac{\\sum d_{x_i}d_{y_i}}{\\sqrt{\\sum d_{x_i}^2 \\sum d_{y_i}^2}}",
        "$$",
        "",
        "$$",
        (
            f"r = \\frac{{{fmt_br(stats['SOMA dx*dy'])}}}"
            f"{{\\sqrt{{{fmt_br(stats['SOMA dx^2'])}\\cdot {fmt_br(stats['SQTot = SOMA dy^2'])}}}}}"
            f" = {fmt_br(stats['r manual'])}"
        ),
        "$$",
        "",
        "### 4. Reta de regressão",
        "",
        "A inclinação da reta é:",
        "",
        "$$",
        "b = \\frac{\\sum d_{x_i}d_{y_i}}{\\sum d_{x_i}^2}",
        "$$",
        "",
        "$$",
        f"b = \\frac{{{fmt_br(stats['SOMA dx*dy'])}}}{{{fmt_br(stats['SOMA dx^2'])}}} = {fmt_br(stats['b manual'])}",
        "$$",
        "",
        "O intercepto é:",
        "",
        "$$",
        "a = \\bar{y} - b\\bar{x}",
        "$$",
        "",
        "$$",
        f"a = {fmt_br(stats['media y'])} - {fmt_br(stats['b manual'])}\\cdot {fmt_br(stats['media x'])} = {fmt_br(stats['a manual'])}",
        "$$",
        "",
        "Logo, a reta ajustada é:",
        "",
        "$$",
        f"\\hat{{y}} = {fmt_br(stats['a manual'])} + {fmt_br(stats['b manual'])}x",
        "$$",
        "",
        "### 5. Coeficiente de determinação",
        "",
        "$$",
        f"R^2 = r^2 = ({fmt_br(stats['r manual'])})^2 = {fmt_br(stats['R2 manual = r^2'])}",
        "$$",
        "",
        "Também é possível calcular $R^2$ pela decomposição da ANOVA:",
        "",
        "$$",
        f"R^2 = \\frac{{\\mathrm{{SQ}}_{{Reg}}}}{{\\mathrm{{SQ}}_{{Tot}}}} = \\frac{{{fmt_br(stats['Regressao SQ'])}}}{{{fmt_br(stats['Total SQ'])}}} = {fmt_br(stats['R2 pela ANOVA'])}",
        "$$",
        "",
        "### 6. ANOVA da regressão",
        "",
        "A decomposição da variabilidade é:",
        "",
        "$$",
        f"\\mathrm{{SQ}}_{{Tot}} = \\mathrm{{SQ}}_{{Reg}} + \\mathrm{{SQ}}_{{Res}} = {fmt_br(stats['Regressao SQ'])} + {fmt_br(stats['Residuos SQ'])} = {fmt_br(stats['Total SQ'])}",
        "$$",
        "",
        "Os quadrados médios são:",
        "",
        "$$",
        f"\\mathrm{{QM}}_{{Reg}} = \\frac{{\\mathrm{{SQ}}_{{Reg}}}}{{\\mathrm{{gl}}_{{Reg}}}} = \\frac{{{fmt_br(stats['Regressao SQ'])}}}{{{gl_reg}}} = {fmt_br(stats['Regressao QM'])}",
        "$$",
        "",
        "$$",
        f"\\mathrm{{QM}}_{{Res}} = \\frac{{\\mathrm{{SQ}}_{{Res}}}}{{\\mathrm{{gl}}_{{Res}}}} = \\frac{{{fmt_br(stats['Residuos SQ'])}}}{{{gl_res}}} = {fmt_br(stats['Residuos QM'])}",
        "$$",
        "",
        "A estatística $F$ é:",
        "",
        "$$",
        f"F = \\frac{{\\mathrm{{QM}}_{{Reg}}}}{{\\mathrm{{QM}}_{{Res}}}} = \\frac{{{fmt_br(stats['Regressao QM'])}}}{{{fmt_br(stats['Residuos QM'])}}} = {fmt_br(stats['Regressao F'])}",
        "$$",
        "",
        "Resumo dos graus de liberdade:",
        "",
        f"- $\\mathrm{{gl}}_{{Reg}} = {gl_reg}$",
        f"- $\\mathrm{{gl}}_{{Res}} = n - 2 = {n} - 2 = {gl_res}$",
        f"- $\\mathrm{{gl}}_{{Tot}} = n - 1 = {n} - 1 = {gl_tot}$",
        "",
        "Tabela ANOVA final:",
        "",
        "| Fonte | $\\mathrm{SQ}$ | $\\mathrm{gl}$ | $\\mathrm{QM}$ | $F$ |",
        "| --- | ---: | ---: | ---: | ---: |",
        f"| Regressão | {fmt_br(stats['Regressao SQ'])} | {gl_reg} | {fmt_br(stats['Regressao QM'])} | {fmt_br(stats['Regressao F'])} |",
        f"| Resíduos | {fmt_br(stats['Residuos SQ'])} | {gl_res} | {fmt_br(stats['Residuos QM'])} |  |",
        f"| Total | {fmt_br(stats['Total SQ'])} | {gl_tot} |  |  |",
    ]


def definitions_section(ws) -> list[str]:
    x_label = pt_br_text(ws.cell(9, 2).value)
    y_label = pt_br_text(ws.cell(9, 3).value)
    return [
        "### Definições das variáveis e símbolos",
        "",
        "| Símbolo ou termo | Significado |",
        "| --- | --- |",
        f"| `x` | Variável explicativa: {md_escape(x_label)}. |",
        f"| `y` | Variável resposta: {md_escape(y_label)}. |",
        "| `n` | Número de observações usadas no cálculo. |",
        "| `x̄` | Média dos valores de `x`. |",
        "| `ȳ` | Média dos valores de `y`. |",
        "| `dx` | Desvio de cada valor de `x` em relação à média: `x - x̄`. |",
        "| `dy` | Desvio de cada valor de `y` em relação à média: `y - ȳ`. |",
        "| `dx*dy` | Produto dos desvios; mostra como `x` e `y` variam em conjunto. |",
        "| `dx^2` | Quadrado do desvio de `x`; mede a variação de `x`. |",
        "| `dy^2` | Quadrado do desvio de `y`; soma essas parcelas para obter `SQTot`. |",
        "| `r` | Correlação linear de Pearson; mede força e direção da associação linear. |",
        "| `R2` | Coeficiente de determinação; proporção da variação de `y` explicada pela regressão. |",
        "| `a` | Intercepto da reta; valor previsto de `y` quando `x = 0`. |",
        "| `b` | Inclinação da reta; variação prevista em `y` para aumento de 1 unidade em `x`. |",
        "| `ŷ` | Valor previsto pela reta de regressão. |",
        "| `resíduo` | Erro de previsão: valor observado menos valor previsto, isto é, `y - ŷ`. |",
        "| `SQ` | Soma de quadrados; medida de variação acumulada. |",
        "| `SQReg` | Soma de quadrados da regressão; parte da variação de `y` explicada pela reta. |",
        "| `SQRes` | Soma de quadrados dos resíduos; parte da variação de `y` não explicada pela reta. |",
        "| `SQTot` | Soma de quadrados total; variação total observada em `y`. |",
        "| `gl` | Graus de liberdade; quantidade de informação independente usada em cada fonte da ANOVA. |",
        "| `QM` | Quadrado médio; calculado por `SQ / gl`. |",
        "| `F` | Estatística da ANOVA; compara `QMReg` com `QMRes`. |",
    ]


def formulas_section() -> list[str]:
    return [
        "### Fórmulas estatísticas e funções da planilha",
        "",
        "#### Fórmulas estatísticas",
        "",
        "| Medida | Fórmula | Leitura didática |",
        "| --- | --- | --- |",
        "| Média de `x` | `x̄ = SOMA(x) / n` | Valor médio da variável explicativa. |",
        "| Média de `y` | `ȳ = SOMA(y) / n` | Valor médio da variável resposta. |",
        "| Desvio de `x` | `dx = x - x̄` | Distância de cada `x` até sua média. |",
        "| Desvio de `y` | `dy = y - ȳ` | Distância de cada `y` até sua média. |",
        "| Correlação | `r = SOMA(dx*dy) / RAIZ(SOMA(dx^2) * SOMA(dy^2))` | Mede associação linear entre `x` e `y`. |",
        "| Inclinação | `b = SOMA(dx*dy) / SOMA(dx^2)` | Mostra quanto `y` tende a variar quando `x` aumenta 1 unidade. |",
        "| Intercepto | `a = ȳ - b*x̄` | Completa a equação da reta. |",
        "| Reta de regressão | `ŷ = a + b*x` | Gera o valor previsto de `y`. |",
        "| Resíduo | `e = y - ŷ` | Diferença entre valor observado e valor previsto. |",
        "| SQTotal | `SQTot = SOMA((y - ȳ)^2)` | Variação total de `y`. |",
        "| SQResíduos | `SQRes = SOMA((y - ŷ)^2)` | Variação que a reta não explicou. |",
        "| SQRegressão | `SQReg = SOMA((ŷ - ȳ)^2)` | Variação explicada pela reta. |",
        "| R2 pela ANOVA | `R2 = SQReg / SQTot` | Proporção explicada pela regressão. |",
        "| Quadrado médio | `QM = SQ / gl` | Variação média por grau de liberdade. |",
        "| Estatística F | `F = QMReg / QMRes` | Compara variação explicada com variação residual. |",
        "",
        "#### Funções estatísticas do Google Planilhas",
        "",
        "| Função | Para que serve | Modelo da fórmula |",
        "| --- | --- | --- |",
        "| `CORREL` | Calcula a correlação linear de Pearson. | `=CORREL(intervalo_y; intervalo_x)` |",
        "| `SLOPE` | Calcula a inclinação `b` da reta. | `=SLOPE(intervalo_y; intervalo_x)` |",
        "| `INTERCEPT` | Calcula o intercepto `a` da reta. | `=INTERCEPT(intervalo_y; intervalo_x)` |",
        "| `RSQ` | Calcula `R2`. | `=RSQ(intervalo_y; intervalo_x)` |",
        "| `LINEST` | Ajusta a regressão linear aos intervalos informados; o primeiro `TRUE` estima o intercepto e o segundo `TRUE` retorna a matriz completa de estatísticas. | `=LINEST(intervalo_y; intervalo_x; TRUE; TRUE)` |",
    ]


def export_sheet(ws) -> str:
    rows = get_data_rows(ws)
    stats = regression(rows)
    title = ws["A1"].value
    enunciado = ws["A2"].value
    questions = [ws.cell(row, 1).value for row in range(4, 8) if ws.cell(row, 1).value]
    interp_start = find_row(ws, "Interpretacao critica") + 1
    interpretation = lines_until_blank(ws, interp_start)

    out = [
        f"# {title}",
        "",
        "## Enunciado",
        "",
        format_enunciado(enunciado),
        "",
        "## Observação inicial e perguntas orientadoras",
        "",
    ]
    out.extend(f"- {pt_br_text(question).replace(chr(10), ' ')}" for question in questions)
    out.extend(
        [
            "",
            *data_table(ws, rows),
            "",
            *paper_calculation_section(rows, stats),
            "",
            "## Desenvolvimento da solução com o uso de planilha",
            "",
            *definitions_section(ws),
            "",
            *formulas_section(),
            "",
            *summary_table(ws, stats),
            "",
            *anova_table(ws, stats),
            "",
        ]
    )
    out.extend(validation_table(ws, stats))
    out.extend(["", "## Interpretação crítica", ""])
    out.extend(f"- {pt_br_text(line)}" for line in interpretation)
    out.append("")
    return "\n".join(out)


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    for sheet_name in ["Ex 1 e resolucao", "Ex 2 e resolucao"]:
        ws = wb[sheet_name]
        output = EXPORT_DIR / f"Revisao_AV2__{slug(sheet_name)}.md"
        output.write_text(export_sheet(ws), encoding="utf-8")
        print(output.relative_to(ROOT))


if __name__ == "__main__":
    main()
